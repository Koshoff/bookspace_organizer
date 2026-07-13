"""Тестове за счетоводната логика: ДДС групи, сторно нетиране, плащания,
консигнация, сторно журнал и Excel експортите."""


def _paid_sale(db, run_sql, order, items, day, method="В брой (Каса)"):
    db.create_sale(order, "w", items, method)
    sid = [s for s in db.get_sales() if s["order_number"] == order][0]["id"]
    db.set_sale_status(sid, "Платена")
    run_sql("UPDATE sales SET created_at=? WHERE order_number=?", (day, order))
    return sid


def _line(pid, qty, price, cost=10.0, title="X"):
    return {"product_id": pid, "title": title, "quantity": qty,
            "cost_price": cost, "sale_price": price}


def test_vat_breakdown_9_20_0(seed, run_sql):
    db = seed.db
    sid = seed.supplier()
    bk = seed.product("BK", "Книга", sid, cover=21.80, vat=9, fiscal_group="Б")
    mg = seed.product("MG", "Чаша", sid, cover=24.00, vat=20,
                      fiscal_group="В", product_type="Стока")
    seed.deliver(sid, bk, 10, 12.0, doc="D1")
    seed.deliver(sid, mg, 10, 10.0, doc="D2")
    _paid_sale(db, run_sql, "O1", [_line(bk, 1, 21.80), _line(mg, 1, 24.00)],
               "2026-06-15")
    vb = db.get_vat_breakdown("2026-06-01", "2026-06-30")
    assert vb["Б"]["base"] == 20.00 and vb["Б"]["vat"] == 1.80
    assert vb["В"]["base"] == 20.00 and vb["В"]["vat"] == 4.00
    assert vb["Д"]["gross"] == 0.0


def test_vat_journal_has_20_percent_row(seed, run_sql):
    db = seed.db
    sid = seed.supplier()
    mg = seed.product("MG", "Чаша", sid, cover=24.00, vat=20,
                      fiscal_group="В", product_type="Стока")
    seed.deliver(sid, mg, 10, 10.0)
    _paid_sale(db, run_sql, "O1", [_line(mg, 2, 24.00)], "2026-06-15")
    j = db.get_sales_journal("2026-06-01", "2026-06-30")
    rates = {r["ДДС ставка"] for r in j}
    assert "20%" in rates
    row = [r for r in j if r["ДДС ставка"] == "20%"][0]
    assert row["Обща стойност с ДДС"] == 48.0 and row["Фискална група"] == "В"


def test_storno_netting_same_period(seed, run_sql):
    db = seed.db
    sid = seed.supplier()
    mg = seed.product("MG", "Чаша", sid, cover=24.00, vat=20,
                      fiscal_group="В", product_type="Стока")
    seed.deliver(sid, mg, 20, 10.0)
    _paid_sale(db, run_sql, "A", [_line(mg, 1, 24.00)], "2026-06-15")
    sB = _paid_sale(db, run_sql, "B", [_line(mg, 1, 24.00)], "2026-06-10")
    db.cancel_sale(sB, "BONB")
    run_sql("UPDATE credit_notes SET created_at='2026-06-16' WHERE original_receipt='BONB'")
    vb = db.get_vat_breakdown("2026-06-01", "2026-06-30")
    # A остава (24), B се неутрализира от своето КИ → нето 24
    assert vb["В"]["gross"] == 24.0


def test_storno_netting_cross_period(seed, run_sql):
    db = seed.db
    sid = seed.supplier()
    mg = seed.product("MG", "Чаша", sid, cover=24.00, vat=20,
                      fiscal_group="В", product_type="Стока")
    seed.deliver(sid, mg, 20, 10.0)
    _paid_sale(db, run_sql, "A", [_line(mg, 1, 24.00)], "2026-06-15")
    sC = _paid_sale(db, run_sql, "C", [_line(mg, 1, 24.00)], "2026-05-20")
    db.cancel_sale(sC, "BONC")
    run_sql("UPDATE credit_notes SET created_at='2026-06-05' WHERE original_receipt='BONC'")
    vb = db.get_vat_breakdown("2026-06-01", "2026-06-30")
    # A (+24) в юни, C-то КИ (−24) в юни → нето 0
    assert vb["В"]["gross"] == 0.0


def test_journal_and_breakdown_consistent(seed, run_sql):
    db = seed.db
    sid = seed.supplier()
    mg = seed.product("MG", "Чаша", sid, cover=24.00, vat=20,
                      fiscal_group="В", product_type="Стока")
    seed.deliver(sid, mg, 20, 10.0)
    _paid_sale(db, run_sql, "A", [_line(mg, 1, 24.00)], "2026-06-15")
    sB = _paid_sale(db, run_sql, "B", [_line(mg, 1, 24.00)], "2026-06-10")
    db.cancel_sale(sB, "BONB")
    run_sql("UPDATE credit_notes SET created_at='2026-06-16' WHERE original_receipt='BONB'")
    j = db.get_sales_journal("2026-06-01", "2026-06-30")
    net = sum(r["Обща стойност с ДДС"] for r in j)
    vb = db.get_vat_breakdown("2026-06-01", "2026-06-30")
    assert abs(net - vb["В"]["gross"]) < 0.01


def test_payment_breakdown_cash_and_voucher(seed, run_sql):
    db = seed.db
    sid = seed.supplier()
    bk = seed.product("BK", "Книга", sid, cover=20.0)
    seed.deliver(sid, bk, 10, 12.0)
    _paid_sale(db, run_sql, "O1", [_line(bk, 2, 20.0)], "2026-06-15",
               method="В брой (Каса)")
    pay = db.get_sales_payment_breakdown("2026-06-01", "2026-06-30")
    assert abs(pay.get("В брой (Каса)", 0) - 40.0) < 0.01


def test_returns_journal(seed, run_sql):
    db = seed.db
    sid = seed.supplier()
    bk = seed.product("BK", "Книга", sid, cover=20.0)
    seed.deliver(sid, bk, 10, 12.0)
    s1 = _paid_sale(db, run_sql, "O1", [_line(bk, 2, 20.0)], "2026-06-15")
    db.cancel_sale(s1, "BON1")
    run_sql("UPDATE credit_notes SET created_at='2026-06-16' WHERE original_receipt='BON1'")
    ret = db.get_returns_journal("2026-06-01", "2026-06-30")
    assert len(ret) == 1
    assert ret[0]["Върната сума"] == -40.0 and ret[0]["Върнати бройки"] == 2


def test_consignment_report_first_sold_across_time(seed, run_sql):
    db = seed.db
    sid = seed.supplier("Изд. X")
    p = seed.product("K", "Книга", sid, cover=12.0)
    # 10 консигнация + 10 купени
    db.create_delivery(sid, "Протокол консигнация", "C1", "2025-12-01",
                       [{"product_id": p, "quantity": 10,
                         "settlement_type": "Консигнация",
                         "supplier_percent": 35, "delivery_price": 5.0}],
                       "Консигнация (отложено)")
    db.create_delivery(sid, "Фактура", "B1", "2025-12-01",
                       [{"product_id": p, "quantity": 10,
                         "settlement_type": "Купена",
                         "supplier_percent": 35, "delivery_price": 5.0}],
                       "По банка")
    _paid_sale(db, run_sql, "BEFORE", [_line(p, 8, 12.0, cost=5.0)], "2025-01-15")
    _paid_sale(db, run_sql, "INPER", [_line(p, 4, 12.0, cost=5.0)], "2026-06-15")
    rep = db.get_consignment_report("2026-06-01", "2026-06-30")
    # 10 консигнация, 8 продадени преди → остават 2; в периода 4 продадени → 2 консигнационни
    assert len(rep) == 1
    assert rep[0]["sold_qty"] == 2 and abs(rep[0]["owed_to_publisher"] - 10.0) < 0.01


def test_full_accounting_excel_bytes(seed, run_sql):
    db = seed.db
    sid = seed.supplier()
    bk = seed.product("BK", "Книга", sid, cover=21.80)
    seed.deliver(sid, bk, 10, 12.0)
    _paid_sale(db, run_sql, "O1", [_line(bk, 1, 21.80)], "2026-06-15")
    data = db.build_full_accounting_excel("2026-06-01", "2026-06-30")
    assert isinstance(data, (bytes, bytearray)) and data[:2] == b"PK"   # xlsx zip
    # проверяваме, че листовете съществуват
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data))
    assert set(["Обобщение", "Дневник на Продажбите", "Сторно",
                "Консигнация"]).issubset(set(wb.sheetnames))
