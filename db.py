import sqlite3

DB_FILE = "bookspace.db"

def get_connection():
    """
    Връща връзка към базата.
    row_factory = sqlite3.Row значи, че редовете се връщат като речник-подобни
    обекти — достъпваш колоните по ИМЕ (row["name"]), не по индекс (row[0]).
    Това прави кода четим и устойчив на смяна на реда на колоните.
    """
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    # Включваме проверката на външните ключове — пак, защото SQLite по
    # подразбиране НЕ я налага и трябва да се пуска при всяка връзка.
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


# ---------- ДОСТАВЧИЦИ (Модул 1) ----------

def add_supplier(name, bulstat, mol, address, phone, email, default_discount):
    """Добавя нов доставчик. Връща (True, съобщение) или (False, грешка)."""
    conn = get_connection()
    try:
        # Параметризирана заявка с ? — НИКОГА не слепвай стойности в SQL низа
        # с f-string! Това е защита срещу SQL injection и счупване при кавички.
        conn.execute(
            """INSERT INTO suppliers
               (name, bulstat, mol, address, phone, email, default_discount)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (name, bulstat, mol, address, phone, email, default_discount)
        )
        conn.commit()
        return True, "Доставчикът е добавен успешно."
    except sqlite3.IntegrityError:
        # Хваща нарушение на UNIQUE — значи вече има доставчик с това име.
        return False, f"Вече съществува доставчик с име „{name}“."
    finally:
        conn.close()  # винаги затваряме връзката, дори при грешка


def get_all_suppliers():
    """Връща всички доставчици, подредени по име."""
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM suppliers ORDER BY name"
    ).fetchall()
    conn.close()
    return rows


# ---------- ПРОДУКТИ / КАТАЛОГ (Модул 2) ----------

def get_all_products():
    """Връща всички книги ЗАЕДНО с името на доставчика и текущата наличност.
    Това е първата ни заявка с JOIN — обединява две таблици в един резултат."""
    conn = get_connection()
    rows = conn.execute(
        """SELECT
               p.id,
               p.isbn,
               p.title,
               p.author,
               s.name AS supplier_name,        -- името идва от suppliers чрез JOIN
               p.cover_price,
               p.vat_rate,
               p.genre,
               -- Наличността = сборът на движенията. COALESCE прави NULL->0,
               -- защото книга без никакви движения няма редове в stock_movements.
               COALESCE((SELECT SUM(quantity_change)
                         FROM stock_movements m
                         WHERE m.product_id = p.id), 0) AS stock
           FROM products p
           JOIN suppliers s ON s.id = p.supplier_id   -- свързваме по релацията
           ORDER BY p.title"""
    ).fetchall()
    conn.close()
    return rows

# ---------- ДОСТАВКИ (Модул 3) ----------

def create_delivery(supplier_id, doc_type, doc_number, doc_date, items, payment_type, operator="система"):
    """
    Създава ЦЯЛА доставка в една транзакция: капак + редове + складови движения.
    'items' е списък от речници, всеки с: product_id, quantity, settlement_type,
    supplier_percent, delivery_price.

    Принцип: или всичко минава, или нищо. Затова няма commit по средата —
    само един commit накрая, и rollback при всяка грешка.
    """
    conn = get_connection()
    try:
        cur = conn.cursor()

        # --- 1. КАПАКЪТ: редът в deliveries ---
        cur.execute(
            """INSERT INTO deliveries (supplier_id, doc_type, doc_number, doc_date, payment_type)
               VALUES (?, ?, ?, ?, ?)""",
            (supplier_id, doc_type, doc_number, doc_date, payment_type)
        )
        # lastrowid ни дава id-то на МЕЖДУ САМО ВКАРАНИЯ ред — трябва ни,
        # за да вържем редовете и движенията към ТАЗИ доставка.
        delivery_id = cur.lastrowid

        # --- 2. РЕДОВЕТЕ + ДВИЖЕНИЯТА за всяка книга ---
        for item in items:
            # 2а. Ред в delivery_items
            cur.execute(
                """INSERT INTO delivery_items
                   (delivery_id, product_id, quantity, settlement_type,
                    supplier_percent, delivery_price)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (delivery_id, item["product_id"], item["quantity"],
                 item["settlement_type"], item["supplier_percent"],
                 item["delivery_price"])
            )

            # 2б. Движение на склада: +количество (входяща стока).
            # ЕТО КЪДЕ наличността реално се вдига над нулата.
            cur.execute(
                """INSERT INTO stock_movements
                   (product_id, movement_type, quantity_change, document_ref, operator)
                   VALUES (?, 'Доставка', ?, ?, ?)""",
                (item["product_id"], item["quantity"],
                 f"{doc_type} №{doc_number}", operator)
            )

        # --- 3. ЕДИН commit за всичко наведнъж ---
        conn.commit()
        return True, f"Доставка №{doc_number} е записана успешно."

    except Exception as e:
        # Нещо се счупи по средата → връщаме базата в изходно положение.
        conn.rollback()
        return False, f"Грешка при записа: {e}"
    finally:
        conn.close()


def get_products_for_delivery():
    """Връща книгите във вид, удобен за избор при доставка:
    {ISBN: {id, title, author, supplier_id}}. Ползва се за сканиране по ISBN."""
    conn = get_connection()
    rows = conn.execute(
        "SELECT id, isbn, title, author, supplier_id FROM products"
    ).fetchall()
    conn.close()
    return {r["isbn"]: dict(r) for r in rows}


def get_deliveries(supplier_id=None, payment_status=None, date_from=None, date_to=None, payment_type=None):
    """
    Връща доставки (капаците) с име на доставчик и обща сума, с опционални филтри.
    Всеки филтър е None, ако не е зададен — тогава не стеснява.
    Това е модел, който ще повтаряме: динамично сглобена WHERE клауза.
    """
    conn = get_connection()

    # Започваме с базова заявка и трупаме условия според подадените филтри.
    query = """
        SELECT
            d.id,
            d.doc_type,
            d.doc_number,
            d.doc_date,
            d.payment_status,
            d.delivery_paid_date,
            s.name AS supplier_name,
            d.payment_type,
            COALESCE(SUM(di.quantity * di.delivery_price), 0) AS total_amount
        FROM deliveries d
        JOIN suppliers s ON s.id = d.supplier_id
        LEFT JOIN delivery_items di ON di.delivery_id = d.id
    """
    conditions = []   # тук трупаме условията
    params = []       # и съответните стойности (пак параметризирано, не слепено!)

    if supplier_id is not None:
        conditions.append("d.supplier_id = ?")
        params.append(supplier_id)
    if payment_status is not None:
        conditions.append("d.payment_status = ?")
        params.append(payment_status)
    if date_from is not None:
        conditions.append("d.doc_date >= ?")
        params.append(date_from)
    if date_to is not None:
        conditions.append("d.doc_date <= ?")
        params.append(date_to)
    if payment_type is not None:
        conditions.append("d.payment_type = ?")
        params.append(payment_type)

    # Ако има условия, лепим ги с WHERE ... AND ...
    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    # GROUP BY е нужен заради SUM — групираме редовете по доставка.
    query += " GROUP BY d.id ORDER BY d.doc_date DESC"

    rows = conn.execute(query, params).fetchall()
    conn.close()
    return rows


def mark_delivery_paid(delivery_id):
    """Маркира доставка като платена И записва точния момент в delivery_paid_date.
    Това е изискването ти за автоматичния timestamp."""
    conn = get_connection()
    try:
        conn.execute(
            """UPDATE deliveries
               SET payment_status = 'Платена',
                   delivery_paid_date = datetime('now', 'localtime')
               WHERE id = ?""",
            (delivery_id,)
        )
        conn.commit()
        return True
    finally:
        conn.close()


def get_delivery_items(delivery_id):
    """Връща книгите в конкретна доставка — за детайлния изглед при клик."""
    conn = get_connection()
    rows = conn.execute(
        """SELECT p.isbn, p.title, di.quantity, di.settlement_type,
                  di.supplier_percent, di.delivery_price
           FROM delivery_items di
           JOIN products p ON p.id = di.product_id
           WHERE di.delivery_id = ?""",
        (delivery_id,)
    ).fetchall()
    conn.close()
    return rows

# ---------- ПРОДАЖБИ / ПОС (Модул 4) ----------

def get_current_stock(product_id):
    """Връща текущата наличност на една книга — сборът на движенията ѝ."""
    conn = get_connection()
    row = conn.execute(
        """SELECT COALESCE(SUM(quantity_change), 0) AS stock
           FROM stock_movements WHERE product_id = ?""",
        (product_id,)
    ).fetchone()
    conn.close()
    return row["stock"]


def create_sale(order_number, waybill_number, items, payment_method, invoice_data=None, operator="система"):
    """
    Създава цяла продажба в една транзакция: капак + редове + движения (с МИНУС).
    Проверява наличност ПРЕДИ да продаде — ако не стига, отказва всичко.

    'items' е списък от речници с: product_id, title, quantity, cost_price, sale_price.
    'invoice_data' е None или речник с данните за фактура.
    """
    conn = get_connection()
    try:
        cur = conn.cursor()

        # --- ПРОВЕРКА НА НАЛИЧНОСТ за всяка книга, ПРЕДИ да пипаме нещо ---
        for item in items:
            row = cur.execute(
                """SELECT COALESCE(SUM(quantity_change), 0) AS stock
                   FROM stock_movements WHERE product_id = ?""",
                (item["product_id"],)
            ).fetchone()
            if row["stock"] < item["quantity"]:
                # Не стига наличност → отказваме ЦЯЛАТА продажба.
                conn.rollback()
                return False, (f"Недостатъчна наличност за „{item['title']}“ "
                               f"(налични: {row['stock']}, искани: {item['quantity']}).")

        # --- 1. КАПАКЪТ ---
        if invoice_data:
            cur.execute(
                """INSERT INTO sales
                   (order_number, waybill_number, payment_method, invoice_issued,
                    invoice_number, buyer_company, buyer_eik, buyer_mol,
                    buyer_address, buyer_email)
                   VALUES (?, ?, ?, 1, ?, ?, ?, ?, ?, ?)""",
                (order_number, waybill_number, payment_method,
                 invoice_data["invoice_number"], invoice_data["buyer_company"],
                 invoice_data["buyer_eik"], invoice_data["buyer_mol"],
                 invoice_data["buyer_address"], invoice_data["buyer_email"])
            )
        else:
            cur.execute(
                """INSERT INTO sales (order_number, waybill_number, payment_method)
                   VALUES (?, ?, ?)""",
                (order_number, waybill_number, payment_method)
            )
        sale_id = cur.lastrowid

        # --- 2. РЕДОВЕТЕ + ДВИЖЕНИЯТА (с МИНУС) ---
        for item in items:
            cur.execute(
                """INSERT INTO sale_items
                   (sale_id, product_id, quantity, cost_price, sale_price)
                   VALUES (?, ?, ?, ?, ?)""",
                (sale_id, item["product_id"], item["quantity"],
                 item["cost_price"], item["sale_price"])
            )
            # МИНУС количество — изходяща стока.
            cur.execute(
                """INSERT INTO stock_movements
                   (product_id, movement_type, quantity_change, document_ref, operator)
                   VALUES (?, 'Продажба', ?, ?, ?)""",
                (item["product_id"], -item["quantity"],
                 f"Поръчка №{order_number}", operator)
            )

        conn.commit()
        return True, f"Продажба №{order_number} е записана успешно."

    except Exception as e:
        conn.rollback()
        return False, f"Грешка при записа: {e}"
    finally:
        conn.close()


def get_product_for_sale(isbn):
    """Връща данните на книга по ISBN за продажба, ВКЛ. доставна цена и наличност.
    Доставната цена я взимаме от последната доставка на тази книга."""
    conn = get_connection()
    row = conn.execute(
        """SELECT p.id, p.isbn, p.title, p.author, p.cover_price,
                  s.name AS supplier_name,
                  COALESCE((SELECT SUM(quantity_change) FROM stock_movements m
                            WHERE m.product_id = p.id), 0) AS stock,
                  COALESCE((SELECT di.delivery_price FROM delivery_items di
                            WHERE di.product_id = p.id
                            ORDER BY di.id DESC LIMIT 1), 0) AS last_cost
           FROM products p
           JOIN suppliers s ON s.id = p.supplier_id
           WHERE p.isbn = ?""",
        (isbn,)
    ).fetchone()
    conn.close()
    return dict(row) if row else None


# ---------- ЖУРНАЛ НА ПРОДАЖБИТЕ (Модул 5) ----------

def get_sales(status=None, date_from=None, date_to=None):
    """
    Връща продажбите с обобщени доставна/продажна суми, с опционални филтри.
    Печалбата се смята после, в Python, от двете суми.
    Същият модел на динамична WHERE клауза като при доставките.
    """
    conn = get_connection()
    query = """
        SELECT
            s.id,
            s.created_at,
            s.order_number,
            s.waybill_number,
            s.status,
            s.payment_date,
            s.payment_method,
            s.supplementary_payment_method,
            s.supplementary_amount,
            s.invoice_issued,
            COALESCE(SUM(si.quantity * si.cost_price), 0) AS total_cost,
            COALESCE(SUM(si.quantity * si.sale_price), 0) AS total_sale
        FROM sales s
        LEFT JOIN sale_items si ON si.sale_id = s.id
    """
    conditions, params = [], []

    if status is not None:
        conditions.append("s.status = ?")
        params.append(status)
    if date_from is not None:
        # Филтрираме по created_at (датата на самата продажба).
        # date() отрязва часа, за да сравняваме само по ден.
        conditions.append("date(s.created_at) >= ?")
        params.append(date_from)
    if date_to is not None:
        conditions.append("date(s.created_at) <= ?")
        params.append(date_to)

    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " GROUP BY s.id ORDER BY s.created_at DESC"

    rows = conn.execute(query, params).fetchall()
    conn.close()
    return rows


def set_sale_status(sale_id, new_status):
    """Сменя статуса на продажба. При 'Платена' записва timestamp в payment_date."""
    conn = get_connection()
    try:
        if new_status == "Платена":
            conn.execute(
                """UPDATE sales
                   SET status = 'Платена',
                       payment_date = datetime('now', 'localtime')
                   WHERE id = ?""",
                (sale_id,)
            )
        else:
            # При други статуси изчистваме payment_date (вече не е платена).
            conn.execute(
                "UPDATE sales SET status = ?, payment_date = NULL WHERE id = ?",
                (new_status, sale_id)
            )
        conn.commit()
        return True
    finally:
        conn.close()


def get_sold_books_for_reorder(date_from, date_to):
    """За експорта: всички продадени книги в периода, групирани по доставчик.
    Сумира количествата на една и съща книга през всички продажби в периода."""
    conn = get_connection()
    rows = conn.execute(
        """SELECT
               sup.name AS supplier_name,
               p.isbn,
               p.title,
               p.author,
               SUM(si.quantity) AS total_sold
           FROM sale_items si
           JOIN sales s   ON s.id = si.sale_id
           JOIN products p ON p.id = si.product_id
           JOIN suppliers sup ON sup.id = p.supplier_id
           WHERE date(s.created_at) >= ? AND date(s.created_at) <= ?
           GROUP BY p.id
           ORDER BY sup.name, p.title""",
        (date_from, date_to)
    ).fetchall()
    conn.close()
    return rows



# ---------- КРЕДИТНИ ИЗВЕСТИЯ / СТОРНО (Модул 6) ----------

def cancel_sale(sale_id, original_receipt, operator="система"):
    """
    Отказва продажба в една транзакция:
    1) сменя статуса на 'Отказана',
    2) връща всяка книга на склад (движение 'Сторно' с ПЛЮС),
    3) записва кредитно известие с номера на оригиналната касова бележка.
    """
    conn = get_connection()
    try:
        cur = conn.cursor()

        # Защита: продажби с ваучер НЕ подлежат на сторно.
        # Политика на бизнеса (ваучерите не се връщат) И техническо ограничение
        # (CHECK ограничението прави складовото движение за ваучер невъзможно).
        voucher_row = cur.execute(
            """SELECT COUNT(*) AS c FROM sale_items
               WHERE sale_id = ? AND voucher_id IS NOT NULL""",
            (sale_id,)
        ).fetchone()
        if voucher_row["c"] > 0:
            conn.rollback()
            return False, ("Тази продажба съдържа ваучер. "
                           "Ваучерите не подлежат на връщане/сторниране.")

        # Първо взимаме редовете на продажбата — трябват ни, за да върнем книгите.
        items = cur.execute(
            "SELECT product_id, quantity, sale_price FROM sale_items WHERE sale_id = ?",
            (sale_id,)
        ).fetchall()

        if not items:
            conn.rollback()
            return False, "Продажбата няма редове за връщане."

        # Сборът, който връщаме на клиента (за журнала на сторната).
        returned_amount = sum(i["quantity"] * i["sale_price"] for i in items)

        # 1) Статусът → 'Отказана'
        cur.execute(
            "UPDATE sales SET status = 'Отказана', payment_date = NULL WHERE id = ?",
            (sale_id,)
        )

        # 2) Връщане на всяка книга на склад: движение 'Сторно' с ПЛЮС.
        for item in items:
            cur.execute(
                """INSERT INTO stock_movements
                   (product_id, movement_type, quantity_change, document_ref, operator)
                   VALUES (?, 'Сторно', ?, ?, ?)""",
                (item["product_id"], item["quantity"],   # ПЛЮС — връщаме на склад
                 f"Кредитно известие (бележка №{original_receipt})", operator)
            )

        # 3) Запис на кредитното известие
        cur.execute(
            """INSERT INTO credit_notes (sale_id, original_receipt, returned_amount)
               VALUES (?, ?, ?)""",
            (sale_id, original_receipt, returned_amount)
        )

        conn.commit()
        return True, f"Сторнирана продажба. Върнати на склад: {len(items)} вид(а) книги."

    except Exception as e:
        conn.rollback()
        return False, f"Грешка при сторното: {e}"
    finally:
        conn.close()


def get_credit_notes(year_month=None):
    """
    Връща кредитните известия, опционално филтрирани по месец (формат 'YYYY-MM').
    strftime изважда годината-месец от датата за филтъра.
    """
    conn = get_connection()
    query = """
        SELECT
            cn.id,
            cn.created_at,
            cn.original_receipt,
            cn.returned_amount,
            s.order_number,
            s.id AS sale_id
        FROM credit_notes cn
        JOIN sales s ON s.id = cn.sale_id
    """
    params = []
    if year_month is not None:
        query += " WHERE strftime('%Y-%m', cn.created_at) = ?"
        params.append(year_month)
    query += " ORDER BY cn.created_at DESC"

    rows = conn.execute(query, params).fetchall()
    conn.close()
    return rows


# ---------- СКЛАД И ОДИТ (Модул 7) ----------

def search_stock(search_term=None):
    """
    Връща книгите с текущата им наличност, опционално филтрирани по търсене
    в заглавие/автор/ISBN. Това е бързият преглед на склада.
    """
    conn = get_connection()
    query = """
        SELECT
            p.id,
            p.isbn,
            p.title,
            p.author,
            s.name AS supplier_name,
            p.cover_price,
            COALESCE((SELECT SUM(quantity_change) FROM stock_movements m
                      WHERE m.product_id = p.id), 0) AS stock
        FROM products p
        JOIN suppliers s ON s.id = p.supplier_id
    """
    params = []
    if search_term:
        # LIKE с % от двете страни = "съдържа някъде". Търсим в три полета.
        # Параметризирано — стойността влиза през ?, не се лепи в текста.
        query += """ WHERE p.title LIKE ? OR p.author LIKE ? OR p.isbn LIKE ?"""
        like = f"%{search_term}%"
        params = [like, like, like]
    query += " ORDER BY p.title"

    rows = conn.execute(query, params).fetchall()
    conn.close()
    return rows


def get_product_history(product_id):
    """
    Връща пълната хронология на движенията на една книга, най-новите отгоре.
    Това Е одитът: всеки ред е едно събитие с документ и оператор.
    """
    conn = get_connection()
    rows = conn.execute(
        """SELECT
               created_at,
               movement_type,
               quantity_change,
               document_ref,
               operator
           FROM stock_movements
           WHERE product_id = ?
           ORDER BY created_at DESC, id DESC""",
        (product_id,)
    ).fetchall()
    conn.close()
    return rows


# ---------- ГЛАВНО КОМАНДНО ТАБЛО (Модул 0) ----------

def get_dashboard_data(date_from, date_to, payment_method=None):
    """
    Връща всички данни за командното табло за зададения период.
    date_from / date_to са низове 'YYYY-MM-DD'.

    ВАЖНО за филтрирането по дата:
    - Приходите (платени продажби) се филтрират по payment_date (кога е платено).
    - Разходите (платени доставки) се филтрират по delivery_paid_date (кога сме платили).
    Така и двете страни гледат КОГА реално са минали парите.
    """
    conn = get_connection()

    # --- КАРТА 1: Приходи = платени продажби, по дата на плащане ---
    rev_query = """SELECT COALESCE(SUM(si.quantity * si.sale_price), 0) AS total
                   FROM sales s
                   JOIN sale_items si ON si.sale_id = s.id
                   WHERE s.status = 'Платена'
                     AND date(s.payment_date) >= ? AND date(s.payment_date) <= ?"""
    rev_params = [date_from, date_to]
    if payment_method:
        rev_query += " AND s.payment_method = ?"
        rev_params.append(payment_method)
    revenue = conn.execute(rev_query, rev_params).fetchone()["total"]

    # --- КАРТА 2: Разходи = платени доставки, по дата на плащане ---
    expenses = conn.execute(
        """SELECT COALESCE(SUM(di.quantity * di.delivery_price), 0) AS total
           FROM deliveries d
           JOIN delivery_items di ON di.delivery_id = d.id
           WHERE d.payment_status = 'Платена'
             AND date(d.delivery_paid_date) >= ? AND date(d.delivery_paid_date) <= ?""",
        (date_from, date_to)
    ).fetchone()["total"]

    # --- КАРТА 4: Брой продажби за периода (по дата на създаване, не отказани) ---
    sales_count = conn.execute(
        """SELECT COUNT(*) AS c
           FROM sales
           WHERE date(created_at) >= ? AND date(created_at) <= ?
             AND status != 'Отказана'""",
        (date_from, date_to)
    ).fetchone()["c"]

    # --- ЛЯВА КОЛОНА: Задължения — неплатени доставки (най-новите отгоре) ---
    # Тези НЕ се филтрират по период — дължиш ги СЕГА, независимо кога са вкарани.
    liabilities = conn.execute(
        """SELECT d.id, d.doc_type, d.doc_number, d.doc_date, s.name AS supplier_name,
                  COALESCE(SUM(di.quantity * di.delivery_price), 0) AS amount
           FROM deliveries d
           JOIN suppliers s ON s.id = d.supplier_id
           LEFT JOIN delivery_items di ON di.delivery_id = d.id
           WHERE d.payment_status = 'Неплатена'
           GROUP BY d.id
           ORDER BY d.doc_date DESC"""
    ).fetchall()

    # --- ДЯСНА КОЛОНА: Вземания — продажби, чакащи плащане (наложен платеж) ---
    receivables = conn.execute(
        """SELECT s.id, s.order_number, s.waybill_number, s.created_at,
                  COALESCE(SUM(si.quantity * si.sale_price), 0) AS amount
           FROM sales s
           LEFT JOIN sale_items si ON si.sale_id = s.id
           WHERE s.status = 'Чака плащане'
           GROUP BY s.id
           ORDER BY s.created_at DESC"""
    ).fetchall()

    # --- ДОЛЕН ПАНЕЛ: последни 10 активности (смесено от трите вида) ---
    # UNION ALL слепва три заявки в общ списък. Всяка дава еднакви колони:
    # дата, тип, документ, стойност — за да се подредят заедно по време.
    activities = conn.execute(
        """SELECT * FROM (
               -- Продажби
               SELECT s.created_at AS ts,
                      CASE
                          WHEN s.order_number LIKE 'VOUCHER-%' THEN 'Издаване ваучер'
                          WHEN s.payment_method = 'Ваучер' THEN 'Продажба (с ваучер)'
                          ELSE 'Продажба'
                      END AS type,
                      'Поръчка №' || COALESCE(s.order_number, '-') AS doc,
                      COALESCE((SELECT SUM(si.quantity * si.sale_price)
                                FROM sale_items si WHERE si.sale_id = s.id), 0) AS value
               FROM sales s

               UNION ALL

               -- Доставки
               SELECT d.created_at AS ts, 'Доставка' AS type,
                      d.doc_type || ' №' || d.doc_number AS doc,
                      COALESCE((SELECT SUM(di.quantity * di.delivery_price)
                                FROM delivery_items di WHERE di.delivery_id = d.id), 0) AS value
               FROM deliveries d

               UNION ALL

               -- Кредитни известия
               SELECT cn.created_at AS ts, 'Кредитно известие' AS type,
                      'Бележка №' || cn.original_receipt AS doc,
                      cn.returned_amount AS value
               FROM credit_notes cn
           )
           ORDER BY ts DESC
           LIMIT 10"""
    ).fetchall()

    conn.close()

    return {
        "revenue": revenue,
        "expenses": expenses,
        "profit": revenue - expenses,        # Карта 3: чиста печалба
        "sales_count": sales_count,
        "liabilities": liabilities,
        "receivables": receivables,
        "activities": activities,
    }

# ---------- СЧЕТОВОДЕН ЕКСПОРТ И ДДС (Модул счетоводство) ----------

def get_sales_journal(date_from, date_to):
    """
    Дневник на ПРОДАЖБИТЕ за ДДС.
    - Книги: 9% ДДС, основа = сума/1.09, ДДС = сума - основа.
    - Ваучери (продажба на ваучер): 0% ДДС, основа = сума, ДДС = 0.
      Различаваме по това дали редът има voucher_id вместо product_id.
    - Кредитните известия влизат отделно с минус, по своята дата.
    """
    conn = get_connection()

    # Продажбите без отказани/чакащи. Сега вадим и какво е плащането.
    sales = conn.execute(
        """SELECT
               s.id AS sale_id,
               s.created_at,
               s.order_number,
               s.invoice_number,
               s.payment_method,
               s.supplementary_payment_method,
               s.supplementary_amount
           FROM sales s
           WHERE s.status = 'Платена'
             AND date(s.created_at) >= ? AND date(s.created_at) <= ?
           ORDER BY s.created_at""",
        (date_from, date_to)
    ).fetchall()

    journal = []

    # За всяка продажба теглим редовете отделно — за да различим ваучерни от книжни.
    for s in sales:
        rows = conn.execute(
            """SELECT product_id, voucher_id, quantity, sale_price
               FROM sale_items WHERE sale_id = ?""",
            (s["sale_id"],)
        ).fetchall()

        # Разделяме на две групи суми: книжна (9%) и ваучерна (0%).
        book_total = 0.0
        voucher_total = 0.0
        for r in rows:
            line_total = r["quantity"] * r["sale_price"]
            if r["voucher_id"] is not None:
                # Ред-ваучер — 0% ДДС, група Д.
                voucher_total += line_total
            else:
                # Ред-книга — 9% ДДС, група Б.
                book_total += line_total

        # Описание на плащането — ясно за счетоводителя.
        if s["payment_method"] == "Ваучер" and s["supplementary_amount"] > 0:
            pay_desc = (f"Ваучер + {s['supplementary_payment_method']} "
                        f"({s['supplementary_amount']:.2f} лв.)")
        else:
            pay_desc = s["payment_method"]

        # Книжната част (ако има) — отделен ред в дневника с 9% ДДС.
        if book_total > 0:
            base = round(book_total / 1.09, 2)
            vat = round(book_total - base, 2)
            journal.append({
                "Документ": s["invoice_number"] or f"Поръчка №{s['order_number'] or '-'}",
                "Данъчна основа": base,
                "ДДС ставка": "9%",
                "Начислено ДДС": vat,
                "Обща стойност с ДДС": round(book_total, 2),
                "Фискална група": "Б",
                "Тип плащане": pay_desc,
                "Статус": "Продажба (книги)",
            })

        # Ваучерната част (ако има) — отделен ред с 0% ДДС, група Д.
        if voucher_total > 0:
            journal.append({
                "Документ": f"Издаване ваучер: {s['order_number'] or '-'}",
                "Данъчна основа": round(voucher_total, 2),
                "ДДС ставка": "0%",
                "Начислено ДДС": 0.0,
                "Обща стойност с ДДС": round(voucher_total, 2),
                "Фискална група": "Д",
                "Тип плащане": pay_desc,
                "Статус": "Продажба (ваучер)",
            })

    # Кредитните известия — по тяхната дата, с минус, както преди.
    credits = conn.execute(
        """SELECT cn.created_at, cn.original_receipt, cn.returned_amount,
                  s.payment_method
           FROM credit_notes cn
           JOIN sales s ON s.id = cn.sale_id
           WHERE date(cn.created_at) >= ? AND date(cn.created_at) <= ?
           ORDER BY cn.created_at""",
        (date_from, date_to)
    ).fetchall()

    for c in credits:
        total = -c["returned_amount"]
        base = round(total / 1.09, 2)
        vat = round(total - base, 2)
        journal.append({
            "Документ": f"КИ към бележка №{c['original_receipt']}",
            "Данъчна основа": base,
            "ДДС ставка": "9%",
            "Начислено ДДС": vat,
            "Обща стойност с ДДС": round(total, 2),
            "Фискална група": "Б",
            "Тип плащане": c["payment_method"],
            "Статус": "Кредитно известие",
        })

    conn.close()
    return journal


def get_purchases_journal(date_from, date_to):
    """
    Дневник на ПОКУПКИТЕ (доставки от издателства), филтриран по дата на ПЛАЩАНЕ.
    Само платените доставки, защото дневникът на покупките отразява платените фактури.
    """
    conn = get_connection()
    rows = conn.execute(
        """SELECT
               d.id,
               d.doc_number,
               d.delivery_paid_date,
               sup.name AS supplier_name,
               COALESCE(SUM(di.quantity * di.delivery_price), 0) AS total_with_vat
           FROM deliveries d
           JOIN suppliers sup ON sup.id = d.supplier_id
           LEFT JOIN delivery_items di ON di.delivery_id = d.id
           WHERE d.payment_status = 'Платена'
             AND date(d.delivery_paid_date) >= ? AND date(d.delivery_paid_date) <= ?
           GROUP BY d.id
           ORDER BY d.delivery_paid_date""",
        (date_from, date_to)
    ).fetchall()
    conn.close()

    journal = []
    for r in rows:
        total = r["total_with_vat"]
        base = round(total / 1.09, 2)
        vat = round(total - base, 2)
        journal.append({
            "Фактура доставчик": r["doc_number"],
            "Доставчик": r["supplier_name"],
            "Данъчна основа": base,
            "ДДС 9%": vat,
            "Общо платена сума": round(total, 2),
        })
    return journal


def get_consignment_report(date_from, date_to):
    """
    Отчет за продадена КОНСИГНАЦИЯ, групиран по издателство.
    Правило: консигнацията се продава ПЪРВА. Затова за всяка книга
    продадените бройки се отчитат като консигнационни до размера на
    реално доставените консигнационни бройки (не повече).

    Работи на ниво "обща бройка за книга", което е достатъчно за отчета
    към издателството (колко им дължим за периода).
    """
    conn = get_connection()
    rows = conn.execute(
        """
        WITH
        -- Колко КОНСИГНАЦИОННИ бройки са доставени за всяка книга (общо, за всички време).
        consigned AS (
            SELECT product_id, SUM(quantity) AS consigned_qty
            FROM delivery_items
            WHERE settlement_type = 'Консигнация'
            GROUP BY product_id
        ),
        -- Колко са ПРОДАДЕНИ за периода (без отказани), за всяка книга.
        sold AS (
            SELECT si.product_id,
                   SUM(si.quantity) AS sold_qty,
                   -- средни цени, за да смятаме дължимо и марж на бройка
                   AVG(si.cost_price) AS avg_cost,
                   AVG(si.sale_price) AS avg_sale
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            WHERE s.status != 'Отказана'
              AND date(s.created_at) >= ? AND date(s.created_at) <= ?
            GROUP BY si.product_id
        )
        SELECT
            sup.name AS supplier_name,
            -- Консигнационни продадени = по-малкото от (продадени, доставени консигнация).
            -- MIN гарантира правилото "не повече от реално доставените консигнационни".
            SUM(MIN(sold.sold_qty, consigned.consigned_qty)) AS sold_qty,
            -- Дължимо към издателството: консигнационни продадени × доставна цена.
            SUM(MIN(sold.sold_qty, consigned.consigned_qty) * sold.avg_cost) AS owed_to_publisher,
            -- Марж на книжарницата върху тези консигнационни бройки.
            SUM(MIN(sold.sold_qty, consigned.consigned_qty) * (sold.avg_sale - sold.avg_cost)) AS bookstore_margin
        FROM sold
        JOIN consigned ON consigned.product_id = sold.product_id
        JOIN products p ON p.id = sold.product_id
        JOIN suppliers sup ON sup.id = p.supplier_id
        GROUP BY sup.id
        ORDER BY sup.name
        """,
        (date_from, date_to)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]



# ---------- ГЕНЕРИРАНЕ НА EXCEL ДНЕВНИК ----------

def build_accounting_excel(date_from, date_to):
    """
    Връща Excel файл (в паметта, като bytes) с два листа:
    'Продажби' и 'Покупки'. Сумите с ДДС са стойности; основата и ДДС
    са Excel ФОРМУЛИ, за да е документът проверим и преизчислим.
    Връща bytes, готови за st.download_button.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from io import BytesIO

    wb = Workbook()

    # --- ЛИСТ 1: ПРОДАЖБИ ---
    ws = wb.active
    ws.title = "Продажби"
    headers = ["Документ", "Данъчна основа", "ДДС ставка", "Начислено ДДС",
               "Обща стойност с ДДС", "Фискална група", "Тип плащане", "Статус"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)   # удебелени заглавия

    sales = get_sales_journal(date_from, date_to)
    # Започваме от ред 2 (ред 1 са заглавията).
    row = 2
    for s in sales:
        total = s["Обща стойност с ДДС"]
        is_voucher = s["ДДС ставка"] == "0%"
        ws.cell(row=row, column=1, value=s["Документ"])
        # При ваучер (0% ДДС) основата = сумата, ДДС = 0. При книги — формули за обратно ДДС.
        if is_voucher:
            ws.cell(row=row, column=2, value=total)         # основа = сума
            ws.cell(row=row, column=3, value="0%")
            ws.cell(row=row, column=4, value=0)              # ДДС = 0
        else:
            ws.cell(row=row, column=2, value=f"=E{row}/1.09")
            ws.cell(row=row, column=3, value=s["ДДС ставка"])
            ws.cell(row=row, column=4, value=f"=E{row}-B{row}")
        ws.cell(row=row, column=5, value=total)
        ws.cell(row=row, column=6, value=s["Фискална група"])
        ws.cell(row=row, column=7, value=s["Тип плащане"])
        ws.cell(row=row, column=8, value=s["Статус"])
        row += 1

    # Ред с тотали (формули, които сумират колоните).
    if row > 2:
        ws.cell(row=row, column=1, value="ОБЩО").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"=SUM(B2:B{row-1})").font = Font(bold=True)
        ws.cell(row=row, column=4, value=f"=SUM(D2:D{row-1})").font = Font(bold=True)
        ws.cell(row=row, column=5, value=f"=SUM(E2:E{row-1})").font = Font(bold=True)

    # --- ЛИСТ 2: ПОКУПКИ ---
    ws2 = wb.create_sheet("Покупки")
    headers2 = ["Фактура доставчик", "Доставчик", "Данъчна основа",
                "ДДС 9%", "Общо платена сума"]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    purchases = get_purchases_journal(date_from, date_to)
    row = 2
    for p in purchases:
        total = p["Общо платена сума"]
        ws2.cell(row=row, column=1, value=p["Фактура доставчик"])
        ws2.cell(row=row, column=2, value=p["Доставчик"])
        ws2.cell(row=row, column=3, value=f"=E{row}/1.09")    # основа
        ws2.cell(row=row, column=4, value=f"=E{row}-C{row}")  # ДДС
        ws2.cell(row=row, column=5, value=total)
        row += 1

    if row > 2:
        ws2.cell(row=row, column=1, value="ОБЩО").font = Font(bold=True)
        ws2.cell(row=row, column=3, value=f"=SUM(C2:C{row-1})").font = Font(bold=True)
        ws2.cell(row=row, column=4, value=f"=SUM(D2:D{row-1})").font = Font(bold=True)
        ws2.cell(row=row, column=5, value=f"=SUM(E2:E{row-1})").font = Font(bold=True)

    # Записваме в паметта (не на диск) и връщаме bytes.
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()    


def get_delivery_payment_breakdown(date_from=None, date_to=None):
    """
    Връща разбивка на доставките по начин на плащане за периода:
    за всеки тип — обща сума и брой документи. Ползва се за счетоводния
    брояч в журнала и за разбивката в таблото.
    """
    conn = get_connection()
    query = """
        SELECT
            d.payment_type,
            COUNT(DISTINCT d.id) AS doc_count,
            COALESCE(SUM(di.quantity * di.delivery_price), 0) AS total
        FROM deliveries d
        LEFT JOIN delivery_items di ON di.delivery_id = d.id
    """
    conditions, params = [], []
    if date_from is not None:
        conditions.append("date(d.doc_date) >= ?")
        params.append(date_from)
    if date_to is not None:
        conditions.append("date(d.doc_date) <= ?")
        params.append(date_to)
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " GROUP BY d.payment_type"

    rows = conn.execute(query, params).fetchall()
    conn.close()
    # Връщаме речник {тип: {count, total}} за лесен достъп.
    return {r["payment_type"]: {"count": r["doc_count"], "total": r["total"]} for r in rows}



# ---------- МЕСЕЧЕН ОТЧЕТ ЗА СТАТУС НА ПЛАЩАНИЯТА ----------

def get_monthly_payment_report(year_month):
    """
    Връща месечен отчет за плащанията. year_month е низ 'YYYY-MM'.
    Дели поръчките на неплатени (чакащи) и платени, за месеца по дата на създаване.
    Връща речник с трите суми и двата списъка.
    """
    conn = get_connection()

    # Неплатени (чакащи плащане) за месеца
    unpaid = conn.execute(
        """SELECT s.created_at, s.order_number, s.waybill_number,
                  s.payment_method,
                  s.supplementary_payment_method,
                  s.supplementary_amount,
                  COALESCE(SUM(si.quantity * si.sale_price), 0) AS amount
           FROM sales s
           LEFT JOIN sale_items si ON si.sale_id = s.id
           WHERE s.status = 'Чака плащане'
             AND strftime('%Y-%m', s.created_at) = ?
           GROUP BY s.id
           ORDER BY s.created_at""",
        (year_month,)
    ).fetchall()

    # Платени за месеца (с дата на плащане)
    paid = conn.execute(
        """SELECT s.created_at, s.order_number, s.waybill_number,
                  s.payment_method, 
                  s.supplementary_payment_method,
                  s.supplementary_amount,
                  s.payment_date,
                  COALESCE(SUM(si.quantity * si.sale_price), 0) AS amount
           FROM sales s
           LEFT JOIN sale_items si ON si.sale_id = s.id
           WHERE s.status = 'Платена'
             AND strftime('%Y-%m', s.created_at) = ?
           GROUP BY s.id
           ORDER BY s.created_at""",
        (year_month,)
    ).fetchall()

    unpaid_total = sum(r["amount"] for r in unpaid)
    paid_total = sum(r["amount"] for r in paid)

    def describe(r):
        d = dict(r)
        if d["payment_method"] == "Ваучер" and d["supplementary_amount"] > 0:
            d["payment_method"] = (f"Ваучер + {d['supplementary_payment_method']} "
                                   f"({d['supplementary_amount']:.2f} лв.)")
        return d

    return {
        "unpaid": [describe(r) for r in unpaid],
        "paid": [describe(r) for r in paid],
        "unpaid_total": unpaid_total,
        "paid_total": paid_total,
        "turnover": unpaid_total + paid_total,
    }

def build_monthly_payment_excel(year_month):
    """Excel с два листа: 'Неплатени' и 'Платени' за месеца."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from io import BytesIO

    data = get_monthly_payment_report(year_month)
    wb = Workbook()

    # Лист 1: Неплатени
    ws = wb.active
    ws.title = "Неплатени"
    ws.append(["Дата/Час", "Поръчка №", "Товарителница",
               "Начин на плащане", "Сума"])
    for c in ws[1]:
        c.font = Font(bold=True)
    row = 2
    for r in data["unpaid"]:
        ws.cell(row=row, column=1, value=r["created_at"])
        ws.cell(row=row, column=2, value=r["order_number"] or "-")
        ws.cell(row=row, column=3, value=r["waybill_number"] or "-")
        ws.cell(row=row, column=4, value=r["payment_method"])
        ws.cell(row=row, column=5, value=round(r["amount"], 2))
        row += 1
    if row > 2:
        ws.cell(row=row, column=1, value="ОБЩО ВИСЯЩИ").font = Font(bold=True)
        ws.cell(row=row, column=5, value=f"=SUM(E2:E{row-1})").font = Font(bold=True)

    # Лист 2: Платени
    ws2 = wb.create_sheet("Платени")
    ws2.append(["Дата/Час", "Поръчка №", "Товарителница",
                "Начин на плащане", "Сума", "Дата на плащане"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    row = 2
    for r in data["paid"]:
        ws2.cell(row=row, column=1, value=r["created_at"])
        ws2.cell(row=row, column=2, value=r["order_number"] or "-")
        ws2.cell(row=row, column=3, value=r["waybill_number"] or "-")
        ws2.cell(row=row, column=4, value=r["payment_method"])
        ws2.cell(row=row, column=5, value=round(r["amount"], 2))
        ws2.cell(row=row, column=6, value=r["payment_date"])
        row += 1
    if row > 2:
        ws2.cell(row=row, column=1, value="ОБЩО СЪБРАНИ").font = Font(bold=True)
        ws2.cell(row=row, column=5, value=f"=SUM(E2:E{row-1})").font = Font(bold=True)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def add_product(isbn, title, author, supplier_id, cover_price, vat_rate,
                year, cover_type, genre, description,
                product_type="Книга", fiscal_group="Б"):
    """Добавя нов артикул (книга или ваучер).
    product_type: 'Книга' или 'Ваучер'.
    fiscal_group: 'Б' (9% ДДС) или 'Д' (0% ДДС, ваучери).
    """
    conn = get_connection()
    try:
        conn.execute(
            """INSERT INTO products
               (isbn, title, author, supplier_id, cover_price, vat_rate,
                year, cover_type, genre, description, product_type, fiscal_group)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (isbn, title, author, supplier_id, cover_price, vat_rate,
             year, cover_type, genre, description, product_type, fiscal_group)
        )
        conn.commit()
        return True, f"{product_type}-ът е добавен успешно."
    except sqlite3.IntegrityError:
        return False, f"Вече съществува артикул с ISBN „{isbn}“."
    finally:
        conn.close()


def issue_voucher(nominal, payment_method="В брой (Каса)"):
    """
    Издава нов ваучер със срок една година: създава продажба с 0% ДДС
    (ваучерът сам по себе си е необлагаем при издаване), генерира уникален
    код и записва ваучера със статус 'Активен'.
    payment_method: как клиентът плаща ВАУЧЕРА (брой/карта/банка).
    """
    from datetime import date, timedelta

    conn = get_connection()
    try:
        cur = conn.cursor()

        year = date.today().year
        row = cur.execute(
            "SELECT COUNT(*) AS c FROM vouchers WHERE code LIKE ?",
            (f"GIFT-{year}-%",)
        ).fetchone()
        code = f"GIFT-{year}-{row['c'] + 1:05d}"
        valid_until = (date.today() + timedelta(days=365)).isoformat()

        cur.execute(
            """INSERT INTO sales (order_number, status, payment_method, payment_date)
               VALUES (?, 'Платена', ?, datetime('now','localtime'))""",
            (f"VOUCHER-{code}", payment_method)
        )
        sale_id = cur.lastrowid

        cur.execute(
            """INSERT INTO vouchers (code, nominal, valid_until, issued_sale_id)
               VALUES (?, ?, ?, ?)""",
            (code, nominal, valid_until, sale_id)
        )
        voucher_id = cur.lastrowid

        cur.execute(
            """INSERT INTO sale_items (sale_id, voucher_id, quantity,
                                       cost_price, sale_price)
               VALUES (?, ?, 1, 0, ?)""",
            (sale_id, voucher_id, nominal)
        )

        conn.commit()
        return True, {"code": code, "nominal": nominal,
                      "valid_until": valid_until, "sale_id": sale_id}

    except Exception as e:
        conn.rollback()
        return False, f"Грешка при издаване на ваучер: {e}"
    finally:
        conn.close()


def get_all_vouchers(status=None):
    """
    Връща всички ваучери, опционално филтрирани по статус.
    Динамично обновява 'Изтекъл' за активни ваучери с изтекъл срок —
    така списъкът винаги показва вярната картина, без отделен крон.
    """
    conn = get_connection()

    # Първо: маркираме изтеклите. Активни ваучери, чийто срок е минал,
    # стават 'Изтекъл'. Това е "ленивата" актуализация — става при четене,
    # не на заден план. За мащаб на книжарница е напълно достатъчно.
    conn.execute(
        """UPDATE vouchers
           SET status = 'Изтекъл'
           WHERE status = 'Активен'
             AND valid_until IS NOT NULL
             AND date(valid_until) < date('now', 'localtime')"""
    )
    conn.commit()

    # После: четем списъка.
    query = "SELECT * FROM vouchers"
    params = []
    if status is not None:
        query += " WHERE status = ?"
        params.append(status)
    query += " ORDER BY issued_at DESC"

    rows = conn.execute(query, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def find_voucher_by_code(code):
    """Намира ваучер по код. Връща речник или None.
    Ползва се в ПОС-а, когато клиентът дава ваучер за плащане."""
    conn = get_connection()
    row = conn.execute(
        "SELECT * FROM vouchers WHERE code = ?", (code.strip(),)
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def validate_voucher_for_use(code):
    """
    Проверява дали ваучер с дадения код може да се ползва.
    Връща (True, voucher_dict) ако да, или (False, грешка) ако не.
    """
    voucher = find_voucher_by_code(code)
    if voucher is None:
        return False, f"Ваучер с код „{code}“ не съществува."
    if voucher["status"] == "Използван":
        return False, f"Ваучер „{code}“ вече е използван на {voucher['used_at']}."
    if voucher["status"] == "Изтекъл":
        return False, f"Ваучер „{code}“ е изтекъл на {voucher['valid_until']}."
    # Допълнителна проверка — макар че get_all_vouchers ленивото обновява,
    # find_voucher_by_code не го прави, затова сверяваме и тук.
    from datetime import date
    if voucher["valid_until"] and voucher["valid_until"] < date.today().isoformat():
        return False, f"Ваучер „{code}“ е изтекъл на {voucher['valid_until']}."
    return True, voucher


def create_sale_with_voucher(order_number, waybill_number, items, voucher_id,
                             supplementary_method=None, invoice_data=None,
                             operator="система"):
    """
    Продажба с ваучер. Ако сумата надвишава номинала, supplementary_method
    се ползва за разликата (Сценарий А — един метод на доплащане).
    Ако сумата е <= номинала, supplementary_method се игнорира.
    """
    conn = get_connection()
    try:
        cur = conn.cursor()

        # Проверка на наличност (както преди).
        for item in items:
            row = cur.execute(
                """SELECT COALESCE(SUM(quantity_change), 0) AS stock
                   FROM stock_movements WHERE product_id = ?""",
                (item["product_id"],)
            ).fetchone()
            if row["stock"] < item["quantity"]:
                conn.rollback()
                return False, (f"Недостатъчна наличност за „{item['title']}“ "
                               f"(налични: {row['stock']}).")

        # Проверка на ваучера в транзакцията (race safety).
        v_row = cur.execute(
            "SELECT * FROM vouchers WHERE id = ? AND status = 'Активен'",
            (voucher_id,)
        ).fetchone()
        if v_row is None:
            conn.rollback()
            return False, "Ваучерът вече не е активен."

        # Изчисляваме сумата и доплащането.
        total_sale = sum(i["quantity"] * i["sale_price"] for i in items)
        nominal = v_row["nominal"]
        supplement_amount = 0.0

        if total_sale > nominal:
            # Доплащане задължително — трябва да е избран метод.
            if not supplementary_method:
                conn.rollback()
                return False, ("Сумата надвишава номинала на ваучера. "
                               "Изберете метод на доплащане.")
            supplement_amount = round(total_sale - nominal, 2)

        # Капак на продажбата. payment_method = 'Ваучер', плюс полетата за
        # доплащане, ако има. Ако няма — supplementary_amount = 0 и метод NULL.
        if invoice_data:
            cur.execute(
                """INSERT INTO sales (order_number, waybill_number, payment_method,
                                      supplementary_payment_method, supplementary_amount,
                                      invoice_issued, invoice_number, buyer_company,
                                      buyer_eik, buyer_mol, buyer_address, buyer_email)
                   VALUES (?, ?, 'Ваучер', ?, ?, 1, ?, ?, ?, ?, ?, ?)""",
                (order_number, waybill_number,
                 supplementary_method if supplement_amount > 0 else None,
                 supplement_amount,
                 invoice_data["invoice_number"], invoice_data["buyer_company"],
                 invoice_data["buyer_eik"], invoice_data["buyer_mol"],
                 invoice_data["buyer_address"], invoice_data["buyer_email"])
            )
        else:
            cur.execute(
                """INSERT INTO sales (order_number, waybill_number, payment_method,
                                      supplementary_payment_method, supplementary_amount)
                   VALUES (?, ?, 'Ваучер', ?, ?)""",
                (order_number, waybill_number,
                 supplementary_method if supplement_amount > 0 else None,
                 supplement_amount)
            )
        sale_id = cur.lastrowid

        # Редове + движения (същото като преди).
        for item in items:
            cur.execute(
                """INSERT INTO sale_items (sale_id, product_id, quantity,
                                           cost_price, sale_price)
                   VALUES (?, ?, ?, ?, ?)""",
                (sale_id, item["product_id"], item["quantity"],
                 item["cost_price"], item["sale_price"])
            )
            cur.execute(
                """INSERT INTO stock_movements (product_id, movement_type,
                                                quantity_change, document_ref, operator)
                   VALUES (?, 'Продажба', ?, ?, ?)""",
                (item["product_id"], -item["quantity"],
                 f"Поръчка №{order_number} (ваучер)", operator)
            )

        # Маркираме ваучера използван.
        cur.execute(
            """UPDATE vouchers
               SET status = 'Използван',
                   used_at = datetime('now', 'localtime'),
                   used_sale_id = ?
               WHERE id = ?""",
            (sale_id, voucher_id)
        )

        conn.commit()
        if supplement_amount > 0:
            return True, (f"Продажба №{order_number} записана. "
                          f"Ваучер: {nominal:.2f} лв. + доплащане "
                          f"{supplement_amount:.2f} лв. с „{supplementary_method}“.")
        else:
            return True, f"Продажба №{order_number} записана (плащане с ваучер)."

    except Exception as e:
        conn.rollback()
        return False, f"Грешка при записа: {e}"
    finally:
        conn.close()  



# ---------- ГОДИШНО ПРИКЛЮЧВАНЕ И ИНВЕНТАРИЗАЦИЯ ----------

def get_inventory_snapshot(as_of_date):
    """
    Връща пълен опис на склада към избрана дата.
    За всяка книга разделя наличността на 'купени' и 'консигнация' по правилото
    "консигнация се продава първа", и пресмята стойностите.

    as_of_date: низ 'YYYY-MM-DD'. Снимка на склада към края на този ден.
    """
    conn = get_connection()

    rows = conn.execute(
        """
        WITH
        -- Общо доставени КУПЕНИ бройки за всяка книга до as_of_date.
        purchased AS (
            SELECT di.product_id, SUM(di.quantity) AS qty
            FROM delivery_items di
            JOIN deliveries d ON d.id = di.delivery_id
            WHERE di.settlement_type = 'Купена'
              AND date(d.doc_date) <= ?
            GROUP BY di.product_id
        ),
        -- Общо доставени КОНСИГНАЦИОННИ бройки за всяка книга до as_of_date.
        consigned AS (
            SELECT di.product_id, SUM(di.quantity) AS qty
            FROM delivery_items di
            JOIN deliveries d ON d.id = di.delivery_id
            WHERE di.settlement_type = 'Консигнация'
              AND date(d.doc_date) <= ?
            GROUP BY di.product_id
        ),
        -- Общо ПРОДАДЕНИ бройки (без отказани) до as_of_date.
        sold AS (
            SELECT si.product_id, SUM(si.quantity) AS qty
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            WHERE s.status != 'Отказана'
              AND si.product_id IS NOT NULL   -- изключва ваучерните редове
              AND date(s.created_at) <= ?
            GROUP BY si.product_id
        ),
        -- Последна доставна цена за всяка книга (за стойностна оценка).
        last_cost AS (
            SELECT product_id, delivery_price
            FROM delivery_items di1
            WHERE id = (SELECT MAX(di2.id) FROM delivery_items di2
                        WHERE di2.product_id = di1.product_id)
        )
        SELECT
            p.isbn,
            p.title,
            p.author,
            sup.name AS supplier_name,
            p.cover_price,
            COALESCE(purchased.qty, 0) AS total_purchased,
            COALESCE(consigned.qty, 0) AS total_consigned,
            COALESCE(sold.qty, 0)      AS total_sold,
            COALESCE(last_cost.delivery_price, 0) AS unit_cost
        FROM products p
        JOIN suppliers sup ON sup.id = p.supplier_id
        LEFT JOIN purchased ON purchased.product_id = p.id
        LEFT JOIN consigned ON consigned.product_id = p.id
        LEFT JOIN sold      ON sold.product_id = p.id
        LEFT JOIN last_cost ON last_cost.product_id = p.id
        WHERE p.product_type = 'Книга'
        ORDER BY sup.name, p.title
        """,
        (as_of_date, as_of_date, as_of_date)
    ).fetchall()

    conn.close()

    # Сега, в Python, прилагаме правилото "консигнация първа".
    # Това е логика, която е по-четима в код, отколкото в SQL.
    snapshot = []
    for r in rows:
        purchased = r["total_purchased"]
        consigned = r["total_consigned"]
        sold = r["total_sold"]

        # Продажбите първо изяждат консигнацията.
        # Ако продадените са повече от консигнацията, остатъкът яде купените.
        consigned_sold = min(sold, consigned)
        purchased_sold = max(0, sold - consigned)

        # Налични в момента:
        purchased_stock = max(0, purchased - purchased_sold)
        consigned_stock = max(0, consigned - consigned_sold)
        total_stock = purchased_stock + consigned_stock

        if total_stock == 0:
            continue   # пропускаме книги без наличност — не са в инвентара

        # Обратно ДДС за единичната доставна цена (тя е с ДДС, помниш).
        unit_cost_with_vat = r["unit_cost"]
        unit_cost_no_vat = round(unit_cost_with_vat / 1.09, 2)

        # Стойности:
        # - "purchased_value" = собствен актив (купени бройки × доставна без ДДС)
        # - "consigned_value" = задбалансово (консигнация × доставна без ДДС)
        # - "potential_revenue" = пазарна стойност на всичко налично (по корична)
        purchased_value = round(purchased_stock * unit_cost_no_vat, 2)
        consigned_value = round(consigned_stock * unit_cost_no_vat, 2)
        potential_revenue = round(total_stock * r["cover_price"], 2)

        snapshot.append({
            "isbn": r["isbn"],
            "title": r["title"],
            "author": r["author"],
            "supplier": r["supplier_name"],
            "purchased_stock": purchased_stock,
            "consigned_stock": consigned_stock,
            "total_stock": total_stock,
            "unit_cost_no_vat": unit_cost_no_vat,
            "purchased_value": purchased_value,
            "consigned_value": consigned_value,
            "cover_price": r["cover_price"],
            "potential_revenue": potential_revenue,
        })

    return snapshot


def get_dead_inventory():
    """
    Връща списък с книги-кандидати за обезценка по ЗКПО.
    Критерии:
      - В склада има налична стока (positive stock).
      - Първата доставка е била преди поне 12 месеца (книгата е имала време).
      - Нула продажби през последните 12 месеца (трайно отсъствие на оборот).
    Връща данни, нужни за предложение към счетоводителя: количество, доставна
    стойност (без ДДС), последна продажба ако е имало.
    """
    conn = get_connection()

    rows = conn.execute(
        """
        WITH
        -- Текуща наличност на всяка книга от движенията.
        stock AS (
            SELECT product_id, SUM(quantity_change) AS qty
            FROM stock_movements
            GROUP BY product_id
            HAVING SUM(quantity_change) > 0
        ),
        -- Дата на първа доставка за всяка книга.
        first_delivery AS (
            SELECT di.product_id, MIN(d.doc_date) AS first_date
            FROM delivery_items di
            JOIN deliveries d ON d.id = di.delivery_id
            GROUP BY di.product_id
        ),
        -- Дата на ПОСЛЕДНА продажба (без отказани).
        last_sale AS (
            SELECT si.product_id, MAX(s.created_at) AS last_date
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            WHERE s.status != 'Отказана'
              AND si.product_id IS NOT NULL
            GROUP BY si.product_id
        ),
        -- Последна доставна цена (за пресмятане на стойността за обезценка).
        last_cost AS (
            SELECT product_id, delivery_price
            FROM delivery_items di1
            WHERE id = (SELECT MAX(di2.id) FROM delivery_items di2
                        WHERE di2.product_id = di1.product_id)
        )
        SELECT
            p.isbn,
            p.title,
            p.author,
            sup.name AS supplier_name,
            stock.qty AS current_stock,
            first_delivery.first_date,
            last_sale.last_date,
            COALESCE(last_cost.delivery_price, 0) AS unit_cost
        FROM products p
        JOIN suppliers sup ON sup.id = p.supplier_id
        JOIN stock ON stock.product_id = p.id          -- има наличност
        JOIN first_delivery ON first_delivery.product_id = p.id
        LEFT JOIN last_sale ON last_sale.product_id = p.id
        LEFT JOIN last_cost ON last_cost.product_id = p.id
        WHERE
            -- Книгата е била доставена преди поне 12 месеца.
            date(first_delivery.first_date) <= date('now', '-12 months', 'localtime')
            -- И няма продажби в последните 12 месеца.
            -- (или изобщо не е продавана, или последната продажба е стара.)
            AND (last_sale.last_date IS NULL
                 OR date(last_sale.last_date) < date('now', '-12 months', 'localtime'))
            AND p.product_type = 'Книга'
        ORDER BY first_delivery.first_date
        """
    ).fetchall()

    conn.close()

    result = []
    for r in rows:
        unit_cost_no_vat = round(r["unit_cost"] / 1.09, 2)
        total_value = round(r["current_stock"] * unit_cost_no_vat, 2)
        result.append({
            "isbn": r["isbn"],
            "title": r["title"],
            "author": r["author"],
            "supplier": r["supplier_name"],
            "stock": r["current_stock"],
            "first_delivery": r["first_date"],
            "last_sale": r["last_date"] or "никога",
            "unit_cost_no_vat": unit_cost_no_vat,
            "total_value": total_value,
        })
    return result