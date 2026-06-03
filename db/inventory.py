from db.connection import get_connection
# само АКО функцията хваща sqlite3.IntegrityError:
import sqlite3



# ---------- ГОДИШНО ПРИКЛЮЧВАНЕ И ИНВЕНТАРИЗАЦИЯ ----------

def get_inventory_snapshot(as_of_date):
    """
    Връща пълен опис на склада към избрана дата.
    За всяка книга разделя наличността на 'купени' и 'консигнация' по правилото
    "консигнация се продава първа", и пресмята стойностите.

    as_of_date: низ 'YYYY-MM-DD'. Снимка на склада към края на този ден.
    """
    conn = get_connection()

    rows = conn.execute(
        """
        WITH
        -- Общо доставени КУПЕНИ бройки за всяка книга до as_of_date.
        purchased AS (
            SELECT di.product_id, SUM(di.quantity) AS qty
            FROM delivery_items di
            JOIN deliveries d ON d.id = di.delivery_id
            WHERE di.settlement_type = 'Купена'
              AND date(d.doc_date) <= ?
            GROUP BY di.product_id
        ),
        -- Общо доставени КОНСИГНАЦИОННИ бройки за всяка книга до as_of_date.
        consigned AS (
            SELECT di.product_id, SUM(di.quantity) AS qty
            FROM delivery_items di
            JOIN deliveries d ON d.id = di.delivery_id
            WHERE di.settlement_type = 'Консигнация'
              AND date(d.doc_date) <= ?
            GROUP BY di.product_id
        ),
        -- Общо ПРОДАДЕНИ бройки (без отказани) до as_of_date.
        sold AS (
            SELECT si.product_id, SUM(si.quantity) AS qty
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            WHERE s.status != 'Отказана'
              AND si.product_id IS NOT NULL   -- изключва ваучерните редове
              AND date(s.created_at) <= ?
            GROUP BY si.product_id
        ),
        -- Последна доставна цена за всяка книга (за стойностна оценка).
        last_cost AS (
            SELECT product_id, delivery_price
            FROM delivery_items di1
            WHERE id = (SELECT MAX(di2.id) FROM delivery_items di2
                        WHERE di2.product_id = di1.product_id)
        )
        SELECT
            p.isbn,
            p.title,
            p.author,
            sup.name AS supplier_name,
            p.cover_price,
            COALESCE(purchased.qty, 0) AS total_purchased,
            COALESCE(consigned.qty, 0) AS total_consigned,
            COALESCE(sold.qty, 0)      AS total_sold,
            COALESCE(last_cost.delivery_price, 0) AS unit_cost
        FROM products p
        JOIN suppliers sup ON sup.id = p.supplier_id
        LEFT JOIN purchased ON purchased.product_id = p.id
        LEFT JOIN consigned ON consigned.product_id = p.id
        LEFT JOIN sold      ON sold.product_id = p.id
        LEFT JOIN last_cost ON last_cost.product_id = p.id
        WHERE p.product_type = 'Книга'
        ORDER BY sup.name, p.title
        """,
        (as_of_date, as_of_date, as_of_date)
    ).fetchall()

    conn.close()

    # Сега, в Python, прилагаме правилото "консигнация първа".
    # Това е логика, която е по-четима в код, отколкото в SQL.
    snapshot = []
    for r in rows:
        purchased = r["total_purchased"]
        consigned = r["total_consigned"]
        sold = r["total_sold"]

        # Продажбите първо изяждат консигнацията.
        # Ако продадените са повече от консигнацията, остатъкът яде купените.
        consigned_sold = min(sold, consigned)
        purchased_sold = max(0, sold - consigned)

        # Налични в момента:
        purchased_stock = max(0, purchased - purchased_sold)
        consigned_stock = max(0, consigned - consigned_sold)
        total_stock = purchased_stock + consigned_stock

        if total_stock == 0:
            continue   # пропускаме книги без наличност — не са в инвентара

        # Обратно ДДС за единичната доставна цена (тя е с ДДС, помниш).
        unit_cost_with_vat = r["unit_cost"]
        unit_cost_no_vat = round(unit_cost_with_vat / 1.09, 2)

        # Стойности:
        # - "purchased_value" = собствен актив (купени бройки × доставна без ДДС)
        # - "consigned_value" = задбалансово (консигнация × доставна без ДДС)
        # - "potential_revenue" = пазарна стойност на всичко налично (по корична)
        purchased_value = round(purchased_stock * unit_cost_no_vat, 2)
        consigned_value = round(consigned_stock * unit_cost_no_vat, 2)
        potential_revenue = round(total_stock * r["cover_price"], 2)

        snapshot.append({
            "isbn": r["isbn"],
            "title": r["title"],
            "author": r["author"],
            "supplier": r["supplier_name"],
            "purchased_stock": purchased_stock,
            "consigned_stock": consigned_stock,
            "total_stock": total_stock,
            "unit_cost_no_vat": unit_cost_no_vat,
            "purchased_value": purchased_value,
            "consigned_value": consigned_value,
            "cover_price": r["cover_price"],
            "potential_revenue": potential_revenue,
        })

    return snapshot


def get_dead_inventory():
    """
    Връща списък с книги-кандидати за обезценка по ЗКПО.
    Критерии:
      - В склада има налична стока (positive stock).
      - Първата доставка е била преди поне 12 месеца (книгата е имала време).
      - Нула продажби през последните 12 месеца (трайно отсъствие на оборот).
    Връща данни, нужни за предложение към счетоводителя: количество, доставна
    стойност (без ДДС), последна продажба ако е имало.
    """
    conn = get_connection()

    rows = conn.execute(
        """
        WITH
        -- Текуща наличност на всяка книга от движенията.
        stock AS (
            SELECT product_id, SUM(quantity_change) AS qty
            FROM stock_movements
            GROUP BY product_id
            HAVING SUM(quantity_change) > 0
        ),
        -- Дата на първа доставка за всяка книга.
        first_delivery AS (
            SELECT di.product_id, MIN(d.doc_date) AS first_date
            FROM delivery_items di
            JOIN deliveries d ON d.id = di.delivery_id
            GROUP BY di.product_id
        ),
        -- Дата на ПОСЛЕДНА продажба (без отказани).
        last_sale AS (
            SELECT si.product_id, MAX(s.created_at) AS last_date
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            WHERE s.status != 'Отказана'
              AND si.product_id IS NOT NULL
            GROUP BY si.product_id
        ),
        -- Последна доставна цена (за пресмятане на стойността за обезценка).
        last_cost AS (
            SELECT product_id, delivery_price
            FROM delivery_items di1
            WHERE id = (SELECT MAX(di2.id) FROM delivery_items di2
                        WHERE di2.product_id = di1.product_id)
        )
        SELECT
            p.isbn,
            p.title,
            p.author,
            sup.name AS supplier_name,
            stock.qty AS current_stock,
            first_delivery.first_date,
            last_sale.last_date,
            COALESCE(last_cost.delivery_price, 0) AS unit_cost
        FROM products p
        JOIN suppliers sup ON sup.id = p.supplier_id
        JOIN stock ON stock.product_id = p.id          -- има наличност
        JOIN first_delivery ON first_delivery.product_id = p.id
        LEFT JOIN last_sale ON last_sale.product_id = p.id
        LEFT JOIN last_cost ON last_cost.product_id = p.id
        WHERE
            -- Книгата е била доставена преди поне 12 месеца.
            date(first_delivery.first_date) <= date('now', '-12 months', 'localtime')
            -- И няма продажби в последните 12 месеца.
            -- (или изобщо не е продавана, или последната продажба е стара.)
            AND (last_sale.last_date IS NULL
                 OR date(last_sale.last_date) < date('now', '-12 months', 'localtime'))
            AND p.product_type = 'Книга'
        ORDER BY first_delivery.first_date
        """
    ).fetchall()

    conn.close()

    result = []
    for r in rows:
        unit_cost_no_vat = round(r["unit_cost"] / 1.09, 2)
        total_value = round(r["current_stock"] * unit_cost_no_vat, 2)
        result.append({
            "isbn": r["isbn"],
            "title": r["title"],
            "author": r["author"],
            "supplier": r["supplier_name"],
            "stock": r["current_stock"],
            "first_delivery": r["first_date"],
            "last_sale": r["last_date"] or "никога",
            "unit_cost_no_vat": unit_cost_no_vat,
            "total_value": total_value,
        })
    return result


def build_inventory_excel(as_of_date):
    """
    Генерира Excel с инвентаризацията към избрана дата.
    Два листа:
      - 'Собствени активи' — купени бройки × доставна без ДДС.
      - 'Задбалансови активи' — консигнационни бройки × доставна без ДДС.
    Една книга със смесена наличност попада и в двата листа с СВОИТЕ части.
    Връща bytes за st.download_button.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from io import BytesIO

    snapshot = get_inventory_snapshot(as_of_date)
    wb = Workbook()

    # --- ЛИСТ 1: Собствени активи (купени) ---
    ws = wb.active
    ws.title = "Собствени активи"
    
    # Заглавие на справката
    ws.cell(row=1, column=1, value=f"Инвентаризационен опис към {as_of_date}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="Собствени дълготрайни активи (купена стока)").font = Font(italic=True)
    
    # Заглавия на колоните
    headers = ["ISBN", "Заглавие", "Автор", "Доставчик",
               "Налични (бр.)", "Ед. доставна (без ДДС)", "Обща стойност (без ДДС)",
               "Корична цена", "Обща пазарна стойност"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True)
    
    row = 5
    for r in snapshot:
        if r["purchased_stock"] <= 0:
            continue  # тази книга няма купени бройки — не е в собствените активи
        ws.cell(row=row, column=1, value=r["isbn"])
        ws.cell(row=row, column=2, value=r["title"])
        ws.cell(row=row, column=3, value=r["author"])
        ws.cell(row=row, column=4, value=r["supplier"])
        ws.cell(row=row, column=5, value=r["purchased_stock"])
        ws.cell(row=row, column=6, value=r["unit_cost_no_vat"])
        # Колона G = бройки × ед. цена; ФОРМУЛА за да е проверимо.
        ws.cell(row=row, column=7, value=f"=E{row}*F{row}")
        ws.cell(row=row, column=8, value=r["cover_price"])
        ws.cell(row=row, column=9, value=f"=E{row}*H{row}")
        row += 1
    
    # Ред със суми
    if row > 5:
        ws.cell(row=row, column=1, value="ОБЩО").font = Font(bold=True)
        ws.cell(row=row, column=7, value=f"=SUM(G5:G{row-1})").font = Font(bold=True)
        ws.cell(row=row, column=9, value=f"=SUM(I5:I{row-1})").font = Font(bold=True)

    # --- ЛИСТ 2: Задбалансови активи (консигнация) ---
    ws2 = wb.create_sheet("Задбалансови активи")
    
    ws2.cell(row=1, column=1, value=f"Инвентаризационен опис към {as_of_date}").font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value="Задбалансови активи (стока на чуждо съхранение — консигнация)").font = Font(italic=True)
    
    headers2 = ["ISBN", "Заглавие", "Автор", "Издателство",
                "Налични (бр.)", "Ед. доставна (без ДДС)", "Дължима стойност",
                "Корична цена", "Пазарна стойност"]
    for col, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True)
    
    row = 5
    for r in snapshot:
        if r["consigned_stock"] <= 0:
            continue
        ws2.cell(row=row, column=1, value=r["isbn"])
        ws2.cell(row=row, column=2, value=r["title"])
        ws2.cell(row=row, column=3, value=r["author"])
        ws2.cell(row=row, column=4, value=r["supplier"])
        ws2.cell(row=row, column=5, value=r["consigned_stock"])
        ws2.cell(row=row, column=6, value=r["unit_cost_no_vat"])
        ws2.cell(row=row, column=7, value=f"=E{row}*F{row}")
        ws2.cell(row=row, column=8, value=r["cover_price"])
        ws2.cell(row=row, column=9, value=f"=E{row}*H{row}")
        row += 1
    
    if row > 5:
        ws2.cell(row=row, column=1, value="ОБЩО").font = Font(bold=True)
        ws2.cell(row=row, column=7, value=f"=SUM(G5:G{row-1})").font = Font(bold=True)
        ws2.cell(row=row, column=9, value=f"=SUM(I5:I{row-1})").font = Font(bold=True)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
