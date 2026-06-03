from db.connection import get_connection
from datetime import date, timedelta

# само АКО функцията хваща sqlite3.IntegrityError:
import sqlite3



def get_stale_inventory(days_threshold=90, min_margin_after_discount=10):
    """
    Намира залежали продукти — наличност > 0 и нула продажби в последните N дни.

    Алгоритъм за препоръчителна отстъпка:
    - Опитваме отстъпки от {20%, 15%, 10%} в този ред (от агресивна към консервативна).
    - За всяка пробваме новата промо цена и пресмятаме маржа върху продажната.
    - Първата отстъпка, при която маржът остава >= min_margin_after_discount, е препоръката.
    - Ако никоя не работи (тоест дори 10% отстъпка свежда маржа под прага),
      продуктът се изключва — не препоръчваме промоция, която е твърде близо до нулева печалба.

    Марж = (продажна − доставна_без_ДДС) / продажна
    Изразен в проценти.
    """
    from datetime import date, timedelta

    conn = get_connection()
    cutoff_date = (date.today() - timedelta(days=days_threshold)).isoformat()

    rows = conn.execute(
        """
        WITH
        stock AS (
            SELECT product_id, SUM(quantity_change) AS qty
            FROM stock_movements
            GROUP BY product_id
            HAVING SUM(quantity_change) > 0
        ),
        last_cost AS (
            SELECT product_id, delivery_price
            FROM delivery_items di1
            WHERE id = (SELECT MAX(di2.id) FROM delivery_items di2
                        WHERE di2.product_id = di1.product_id)
        ),
        recent_sales AS (
            SELECT DISTINCT si.product_id
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            WHERE s.status != 'Отказана'
              AND si.product_id IS NOT NULL
              AND date(s.created_at) >= ?
        )
        SELECT
            p.isbn,
            p.title,
            sup.name AS supplier_name,
            stock.qty AS current_stock,
            COALESCE(last_cost.delivery_price, 0) AS unit_cost,
            p.cover_price
        FROM products p
        JOIN suppliers sup ON sup.id = p.supplier_id
        JOIN stock ON stock.product_id = p.id
        LEFT JOIN last_cost ON last_cost.product_id = p.id
        WHERE p.id NOT IN (SELECT product_id FROM recent_sales)
          AND p.product_type = 'Книга'
        ORDER BY stock.qty DESC
        """,
        (cutoff_date,)
    ).fetchall()

    conn.close()

    discounts = [20, 15, 10]   # пробваме от агресивна към консервативна
    result = []

    for r in rows:
        cover = r["cover_price"]
        cost_no_vat = r["unit_cost"] / 1.09 if r["unit_cost"] else 0
        stock_qty = r["current_stock"]

        # Текущ марж върху продажната (за справка в таблицата)
        current_margin = ((cover - cost_no_vat) / cover * 100) if cover > 0 else 0

        # Търсим максимална отстъпка, при която маржът остава над прага.
        recommended_discount = 0
        new_price = cover
        new_margin = current_margin

        for d in discounts:
            candidate_price = cover * (1 - d / 100)
            # Маржът след отстъпката, върху новата продажна цена.
            candidate_margin = ((candidate_price - cost_no_vat) / candidate_price * 100) \
                               if candidate_price > 0 else -1
            if candidate_margin >= min_margin_after_discount:
                recommended_discount = d
                new_price = round(candidate_price, 2)
                new_margin = candidate_margin
                break

        # Ако нито една отстъпка не работи — пропускаме продукта.
        # Причина: маржът му е твърде тесен, за да издържи дори 10% отстъпка.
        if recommended_discount == 0:
            continue

        potential_revenue = round(new_price * stock_qty, 2)
        potential_profit = round((new_price - cost_no_vat) * stock_qty, 2)

        result.append({
            "isbn": r["isbn"],
            "title": r["title"],
            "supplier": r["supplier_name"],
            "stock": stock_qty,
            "unit_cost_no_vat": round(cost_no_vat, 2),
            "cover_price": cover,
            "current_margin_percent": round(current_margin, 1),
            "discount_percent": recommended_discount,
            "promo_price": new_price,
            "new_margin_percent": round(new_margin, 1),
            "potential_revenue": potential_revenue,
            "potential_profit": potential_profit,
        })

    return result


def build_promotion_excel(stale_items):
    """
    Excel за импорт в онлайн магазина — ISBN и нова промо цена.
    Минимален формат, лесен за parse от която и да е e-commerce платформа.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Промо цени"

    headers = ["ISBN", "Заглавие", "Оригинална цена", "Промо цена", "Отстъпка %"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    for row_idx, item in enumerate(stale_items, start=2):
        ws.cell(row=row_idx, column=1, value=item["isbn"])
        ws.cell(row=row_idx, column=2, value=item["title"])
        ws.cell(row=row_idx, column=3, value=item["cover_price"])
        ws.cell(row=row_idx, column=4, value=item["promo_price"])
        ws.cell(row=row_idx, column=5, value=item["discount_percent"])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()