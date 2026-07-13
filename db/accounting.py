from db.connection import get_connection
# само АКО функцията хваща sqlite3.IntegrityError:
import sqlite3


# ---------- СЧЕТОВОДЕН ЕКСПОРТ И ДДС (Модул счетоводство) ----------

def get_sales_journal(date_from, date_to):
    """
    Дневник на ПРОДАЖБИТЕ за ДДС.
    - Книги: 9% ДДС, основа = сума/1.09, ДДС = сума - основа.
    - Ваучери (продажба на ваучер): 0% ДДС, основа = сума, ДДС = 0.
      Различаваме по това дали редът има voucher_id вместо product_id.
    - Кредитните известия влизат отделно с минус, по своята дата.
    """
    conn = get_connection()

    # Продажбите без отказани/чакащи. Сега вадим и какво е плащането.
    sales = conn.execute(
        """SELECT
               s.id AS sale_id,
               s.created_at,
               s.order_number,
               s.invoice_number,
               s.payment_method,
               s.supplementary_payment_method,
               s.supplementary_amount
           FROM sales s
           WHERE (s.status = 'Платена'
                  OR EXISTS (SELECT 1 FROM credit_notes cn WHERE cn.sale_id = s.id))
             AND date(s.created_at) >= ? AND date(s.created_at) <= ?
           ORDER BY s.created_at""",
        (date_from, date_to)
    ).fetchall()

    journal = []

    # За всяка продажба теглим редовете отделно — за да различим групите по
    # РЕАЛНИЯ vat_rate/fiscal_group: книги (9% Б), стоки (20% В), ваучери (0% Д).
    for s in sales:
        rows = conn.execute(
            """SELECT si.voucher_id, si.quantity, si.sale_price,
                      p.fiscal_group, p.vat_rate
               FROM sale_items si
               LEFT JOIN products p ON p.id = si.product_id
               WHERE si.sale_id = ?""",
            (s["sale_id"],)
        ).fetchall()

        book9 = goods20 = voucher0 = 0.0
        for r in rows:
            line_total = r["quantity"] * r["sale_price"]
            if r["voucher_id"] is not None:
                voucher0 += line_total
            elif (r["vat_rate"] or 0) >= 20 or r["fiscal_group"] == "В":
                goods20 += line_total
            else:
                book9 += line_total

        # Описание на плащането — ясно за счетоводителя.
        if s["payment_method"] == "Ваучер" and s["supplementary_amount"] > 0:
            pay_desc = (f"Ваучер + {s['supplementary_payment_method']} "
                        f"({s['supplementary_amount']:.2f} лв.)")
        else:
            pay_desc = s["payment_method"]

        doc = s["invoice_number"] or f"Поръчка №{s['order_number'] or '-'}"

        if book9 > 0:
            base = round(book9 / 1.09, 2)
            journal.append({
                "Документ": doc, "Данъчна основа": base, "ДДС ставка": "9%",
                "Начислено ДДС": round(book9 - base, 2),
                "Обща стойност с ДДС": round(book9, 2), "Фискална група": "Б",
                "Тип плащане": pay_desc, "Статус": "Продажба (книги)",
            })
        if goods20 > 0:
            base = round(goods20 / 1.20, 2)
            journal.append({
                "Документ": doc, "Данъчна основа": base, "ДДС ставка": "20%",
                "Начислено ДДС": round(goods20 - base, 2),
                "Обща стойност с ДДС": round(goods20, 2), "Фискална група": "В",
                "Тип плащане": pay_desc, "Статус": "Продажба (стоки)",
            })
        if voucher0 > 0:
            journal.append({
                "Документ": f"Издаване ваучер: {s['order_number'] or '-'}",
                "Данъчна основа": round(voucher0, 2), "ДДС ставка": "0%",
                "Начислено ДДС": 0.0, "Обща стойност с ДДС": round(voucher0, 2),
                "Фискална група": "Д", "Тип плащане": pay_desc,
                "Статус": "Продажба (ваучер)",
            })

    # Кредитните известия — по тяхната дата, с минус. Класифицираме върнатата
    # стока по група (9%/20%) от редовете на оригиналната продажба.
    credits = conn.execute(
        """SELECT cn.created_at, cn.original_receipt, cn.sale_id, s.payment_method
           FROM credit_notes cn
           JOIN sales s ON s.id = cn.sale_id
           WHERE date(cn.created_at) >= ? AND date(cn.created_at) <= ?
           ORDER BY cn.created_at""",
        (date_from, date_to)
    ).fetchall()

    for c in credits:
        lines = conn.execute(
            """SELECT si.voucher_id, si.quantity, si.sale_price,
                      p.fiscal_group, p.vat_rate
               FROM sale_items si
               LEFT JOIN products p ON p.id = si.product_id
               WHERE si.sale_id = ?""",
            (c["sale_id"],)
        ).fetchall()
        r9 = r20 = 0.0
        for l in lines:
            amt = l["quantity"] * l["sale_price"]
            if l["voucher_id"] is not None:
                continue
            elif (l["vat_rate"] or 0) >= 20 or l["fiscal_group"] == "В":
                r20 += amt
            else:
                r9 += amt
        if r9 > 0:
            total = -r9
            base = round(total / 1.09, 2)
            journal.append({
                "Документ": f"КИ към бележка №{c['original_receipt']}",
                "Данъчна основа": base, "ДДС ставка": "9%",
                "Начислено ДДС": round(total - base, 2),
                "Обща стойност с ДДС": round(total, 2), "Фискална група": "Б",
                "Тип плащане": c["payment_method"], "Статус": "Кредитно известие",
            })
        if r20 > 0:
            total = -r20
            base = round(total / 1.20, 2)
            journal.append({
                "Документ": f"КИ към бележка №{c['original_receipt']}",
                "Данъчна основа": base, "ДДС ставка": "20%",
                "Начислено ДДС": round(total - base, 2),
                "Обща стойност с ДДС": round(total, 2), "Фискална група": "В",
                "Тип плащане": c["payment_method"], "Статус": "Кредитно известие",
            })

    conn.close()
    return journal


def get_purchases_journal(date_from, date_to):
    """
    Дневник на ПОКУПКИТЕ (доставки от издателства), филтриран по дата на ПЛАЩАНЕ.
    Само платените доставки, защото дневникът на покупките отразява платените фактури.
    """
    conn = get_connection()
    rows = conn.execute(
        """SELECT
               d.id,
               d.doc_number,
               d.delivery_paid_date,
               sup.name AS supplier_name,
               COALESCE(SUM(di.quantity * di.delivery_price), 0) AS total_with_vat
           FROM deliveries d
           JOIN suppliers sup ON sup.id = d.supplier_id
           LEFT JOIN delivery_items di ON di.delivery_id = d.id
           WHERE d.payment_status = 'Платена'
             AND date(d.delivery_paid_date) >= ? AND date(d.delivery_paid_date) <= ?
           GROUP BY d.id
           ORDER BY d.delivery_paid_date""",
        (date_from, date_to)
    ).fetchall()
    conn.close()

    journal = []
    for r in rows:
        total = r["total_with_vat"]
        base = round(total / 1.09, 2)
        vat = round(total - base, 2)
        journal.append({
            "Фактура доставчик": r["doc_number"],
            "Доставчик": r["supplier_name"],
            "Данъчна основа": base,
            "ДДС 9%": vat,
            "Общо платена сума": round(total, 2),
        })
    return journal


# ---------- СЧЕТОВОДНА СЕКЦИЯ (Б/В/Д, плащания, сторно) ----------

def _payment_desc(payment_method, supp_method, supp_amount):
    """Четимо описание на плащането, вкл. ваучер с доплащане."""
    if payment_method == "Ваучер" and (supp_amount or 0) > 0:
        return f"Ваучер + {supp_method} ({supp_amount:.2f} лв.)"
    return payment_method


def get_sales_vat_journal(date_from, date_to):
    """
    Дневник на продажбите с РАЗДЕЛЕНИ ставки по фискална група на всеки ред:
      - Група Б (9%)  — книги,
      - Група В (20%) — стоки/мърчандайз (vat_rate 20 или fiscal_group 'В'),
      - Група Д (0%)  — ваучерни редове.
    Един документ може да съдържа и трите. Връща по един ред на документ с
    данъчни основи и ДДС за 9% и 20% + ваучерната част.
    Използва РЕАЛНИЯ vat_rate/fiscal_group на продукта, не твърдо 9%.
    """
    conn = get_connection()
    sales = conn.execute(
        """SELECT s.id, s.created_at, s.order_number, s.invoice_number,
                  s.payment_method, s.supplementary_payment_method,
                  s.supplementary_amount
           FROM sales s
           WHERE (s.status = 'Платена'
                  OR EXISTS (SELECT 1 FROM credit_notes cn WHERE cn.sale_id = s.id))
             AND date(s.created_at) >= ? AND date(s.created_at) <= ?
           ORDER BY s.created_at""",
        (date_from, date_to)
    ).fetchall()

    out = []
    for s in sales:
        lines = conn.execute(
            """SELECT si.voucher_id, si.quantity, si.sale_price,
                      p.fiscal_group, p.vat_rate
               FROM sale_items si
               LEFT JOIN products p ON p.id = si.product_id
               WHERE si.sale_id = ?""",
            (s["id"],)
        ).fetchall()

        gross9 = gross20 = voucher0 = 0.0
        for l in lines:
            amt = l["quantity"] * l["sale_price"]
            if l["voucher_id"] is not None:
                voucher0 += amt
            elif (l["vat_rate"] or 0) >= 20 or l["fiscal_group"] == "В":
                gross20 += amt
            else:
                gross9 += amt

        base9 = round(gross9 / 1.09, 2)
        vat9 = round(gross9 - base9, 2)
        base20 = round(gross20 / 1.20, 2)
        vat20 = round(gross20 - base20, 2)

        out.append({
            "Документ": s["invoice_number"] or f"Поръчка №{s['order_number'] or '-'}",
            "Дата": s["created_at"],
            "Оборот 9%": base9,
            "ДДС 9%": vat9,
            "Оборот 20%": base20,
            "ДДС 20%": vat20,
            "Ваучер 0%": round(voucher0, 2),
            "Общо с ДДС": round(gross9 + gross20 + voucher0, 2),
            "Тип плащане": _payment_desc(s["payment_method"],
                                         s["supplementary_payment_method"],
                                         s["supplementary_amount"]),
        })
    conn.close()
    return out


def get_vat_breakdown(date_from, date_to):
    """
    Обобщени суми по трите данъчни групи за периода. НЕТИРА кредитните известия
    (сторно) — както дневникът на продажбите — за да не се над-декларира ДДС.
    Връща {'Б':{base,vat,gross}, 'В':{...}, 'Д':{base,vat,gross}}.

    Забележка: следва конвенцията на get_sales_journal (платени продажби минус
    кредитни известия по дата на КИ). Продажба, платена и сторнирана в ЕДИН и
    същ период, се третира като нетно намаление — същото като в експорта.
    """
    conn = get_connection()
    sale_lines = conn.execute(
        """SELECT si.voucher_id, si.quantity, si.sale_price,
                  p.fiscal_group, p.vat_rate
           FROM sales s
           JOIN sale_items si ON si.sale_id = s.id
           LEFT JOIN products p ON p.id = si.product_id
           WHERE (s.status = 'Платена'
                  OR EXISTS (SELECT 1 FROM credit_notes cn WHERE cn.sale_id = s.id))
             AND date(s.created_at) >= ? AND date(s.created_at) <= ?""",
        (date_from, date_to)
    ).fetchall()
    # Всяко кредитно известие в периода се вади (по своята дата). Оригиналната
    # продажба вече е броена като положителна (виж EXISTS горе), затова
    # изваждането нетира вярно — и за същопериодно, и за минало сторно.
    return_lines = conn.execute(
        """SELECT si.voucher_id, si.quantity, si.sale_price,
                  p.fiscal_group, p.vat_rate
           FROM credit_notes cn
           JOIN sale_items si ON si.sale_id = cn.sale_id
           LEFT JOIN products p ON p.id = si.product_id
           WHERE date(cn.created_at) >= ? AND date(cn.created_at) <= ?""",
        (date_from, date_to)
    ).fetchall()
    conn.close()

    def _cls(r):
        if r["voucher_id"] is not None:
            return "Д"
        if (r["vat_rate"] or 0) >= 20 or r["fiscal_group"] == "В":
            return "В"
        return "Б"

    gross = {"Б": 0.0, "В": 0.0, "Д": 0.0}
    for r in sale_lines:
        gross[_cls(r)] += r["quantity"] * r["sale_price"]
    for r in return_lines:
        gross[_cls(r)] -= r["quantity"] * r["sale_price"]

    def _grp(g, rate):
        if rate == 0:
            return {"base": round(g, 2), "vat": 0.0, "gross": round(g, 2)}
        base = round(g / (1 + rate / 100), 2)
        return {"base": base, "vat": round(g - base, 2), "gross": round(g, 2)}

    return {"Б": _grp(gross["Б"], 9), "В": _grp(gross["В"], 20),
            "Д": _grp(gross["Д"], 0)}


def get_sales_payment_breakdown(date_from, date_to):
    """
    Оборот от платените продажби, разбит по начин на плащане (за съпоставка с
    банка/Z-отчети). Ваучерната част отива в „Ваучер", а доплащането — към
    съответния метод. Връща {метод: сума}.
    """
    from collections import defaultdict
    conn = get_connection()
    rows = conn.execute(
        """SELECT s.payment_method, s.supplementary_payment_method,
                  s.supplementary_amount,
                  COALESCE(SUM(si.quantity * si.sale_price), 0) AS total
           FROM sales s
           LEFT JOIN sale_items si ON si.sale_id = s.id
           WHERE s.status = 'Платена'
             AND date(s.created_at) >= ? AND date(s.created_at) <= ?
           GROUP BY s.id""",
        (date_from, date_to)
    ).fetchall()
    conn.close()

    buckets = defaultdict(float)
    for r in rows:
        total = r["total"]
        pm = r["payment_method"]
        supp = r["supplementary_amount"] or 0
        if pm == "Ваучер":
            buckets["Ваучер"] += max(0.0, total - supp)
            if supp > 0 and r["supplementary_payment_method"]:
                buckets[r["supplementary_payment_method"]] += supp
        else:
            buckets[pm] += total
    return {k: round(val, 2) for k, val in buckets.items()}


def get_returns_journal(date_from, date_to):
    """
    Журнал на сторно операциите (кредитни известия) за периода: дата, оригинална
    бележка, номер поръчка, върната сума (отрицателна) и върнати бройки на склад.
    """
    conn = get_connection()
    rows = conn.execute(
        """SELECT cn.created_at, cn.original_receipt, cn.returned_amount,
                  s.order_number,
                  COALESCE((SELECT SUM(si.quantity) FROM sale_items si
                            WHERE si.sale_id = cn.sale_id
                              AND si.product_id IS NOT NULL), 0) AS returned_units
           FROM credit_notes cn
           JOIN sales s ON s.id = cn.sale_id
           WHERE date(cn.created_at) >= ? AND date(cn.created_at) <= ?
           ORDER BY cn.created_at DESC""",
        (date_from, date_to)
    ).fetchall()
    conn.close()
    result = []
    for r in rows:
        result.append({
            "Дата на сторно": r["created_at"],
            "Оригинална бележка/поръчка": r["original_receipt"]
            or (r["order_number"] or "-"),
            "Върната сума": -round(r["returned_amount"], 2),
            "Върнати бройки": r["returned_units"],
        })
    return result


def build_full_accounting_excel(date_from, date_to):
    """
    Единен счетоводен Excel за периода в стил Sleek Monochrome:
      Лист 1 „Обобщение"  — плащания + ДДС по групи Б/В/Д.
      Лист 2 „Дневник на Продажбите" — по документ с отделни колони
             Оборот 9% / ДДС 9% / Оборот 20% / ДДС 20% (+ Ваучер 0%), SUM тотали.
      Лист 3 „Сторно"     — кредитни известия (върнати суми/бройки).
      Лист 4 „Консигнация" — дължимо към издателствата за периода.
    Тъмни хедъри (#1A1A1A), зебра ефект и =SUM() формули за тоталите.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    HEAD_FILL = PatternFill("solid", fgColor="1A1A1A")
    HEAD_FONT = Font(bold=True, color="FFFFFF")
    ZEBRA = PatternFill("solid", fgColor="F2F2F2")
    BOLD = Font(bold=True)

    def write_header(ws, headers, row=1):
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = HEAD_FILL
            c.font = HEAD_FONT
            c.alignment = Alignment(horizontal="center")

    def zebra_row(ws, row, ncols):
        if row % 2 == 0:
            for col in range(1, ncols + 1):
                ws.cell(row=row, column=col).fill = ZEBRA

    wb = Workbook()

    # ---------- ЛИСТ 1: ОБОБЩЕНИЕ ----------
    ws = wb.active
    ws.title = "Обобщение"
    ws.cell(row=1, column=1, value=f"Счетоводно обобщение {date_from} — {date_to}").font = Font(bold=True, size=14)

    ws.cell(row=3, column=1, value="Оборот по начин на плащане").font = BOLD
    write_header(ws, ["Начин на плащане", "Сума (лв.)"], row=4)
    pay = get_sales_payment_breakdown(date_from, date_to)
    r = 5
    for method, amount in sorted(pay.items()):
        ws.cell(row=r, column=1, value=method)
        ws.cell(row=r, column=2, value=round(amount, 2))
        zebra_row(ws, r, 2)
        r += 1
    if r > 5:
        ws.cell(row=r, column=1, value="ОБЩО").font = BOLD
        ws.cell(row=r, column=2, value=f"=SUM(B5:B{r-1})").font = BOLD
    r += 2

    ws.cell(row=r, column=1, value="ДДС разбивка по данъчни групи").font = BOLD
    r += 1
    write_header(ws, ["Група", "Оборот (основа)", "ДДС", "Общо с ДДС"], row=r)
    vat = get_vat_breakdown(date_from, date_to)
    labels = {"Б": "Група Б (9% — книги)", "В": "Група В (20% — стоки)",
              "Д": "Група Д (0% — ваучери)"}
    r += 1
    for key in ("Б", "В", "Д"):
        g = vat[key]
        ws.cell(row=r, column=1, value=labels[key])
        ws.cell(row=r, column=2, value=g["base"])
        ws.cell(row=r, column=3, value=g["vat"])
        ws.cell(row=r, column=4, value=g["gross"])
        zebra_row(ws, r, 4)
        r += 1
    ws.cell(row=r, column=1, value="ОБЩО").font = BOLD
    ws.cell(row=r, column=2, value=f"=SUM(B{r-3}:B{r-1})").font = BOLD
    ws.cell(row=r, column=3, value=f"=SUM(C{r-3}:C{r-1})").font = BOLD
    ws.cell(row=r, column=4, value=f"=SUM(D{r-3}:D{r-1})").font = BOLD

    # ---------- ЛИСТ 2: ДНЕВНИК НА ПРОДАЖБИТЕ ----------
    ws2 = wb.create_sheet("Дневник на Продажбите")
    headers2 = ["Документ", "Дата", "Оборот 9%", "ДДС 9%", "Оборот 20%",
                "ДДС 20%", "Ваучер 0%", "Общо с ДДС", "Тип плащане"]
    write_header(ws2, headers2)
    journal = get_sales_vat_journal(date_from, date_to)
    row = 2
    for j in journal:
        ws2.cell(row=row, column=1, value=j["Документ"])
        ws2.cell(row=row, column=2, value=j["Дата"])
        ws2.cell(row=row, column=3, value=j["Оборот 9%"])
        ws2.cell(row=row, column=4, value=j["ДДС 9%"])
        ws2.cell(row=row, column=5, value=j["Оборот 20%"])
        ws2.cell(row=row, column=6, value=j["ДДС 20%"])
        ws2.cell(row=row, column=7, value=j["Ваучер 0%"])
        ws2.cell(row=row, column=8, value=j["Общо с ДДС"])
        ws2.cell(row=row, column=9, value=j["Тип плащане"])
        zebra_row(ws2, row, len(headers2))
        row += 1
    if row > 2:
        ws2.cell(row=row, column=1, value="ОБЩО").font = BOLD
        for col in range(3, 9):     # C..H са числови
            letter = chr(ord('A') + col - 1)
            ws2.cell(row=row, column=col,
                     value=f"=SUM({letter}2:{letter}{row-1})").font = BOLD

    # ---------- ЛИСТ 3: СТОРНО ----------
    ws3 = wb.create_sheet("Сторно")
    headers3 = ["Дата на сторно", "Оригинална бележка/поръчка",
                "Върната сума", "Върнати бройки"]
    write_header(ws3, headers3)
    returns = get_returns_journal(date_from, date_to)
    row = 2
    for rr in returns:
        ws3.cell(row=row, column=1, value=rr["Дата на сторно"])
        ws3.cell(row=row, column=2, value=rr["Оригинална бележка/поръчка"])
        ws3.cell(row=row, column=3, value=rr["Върната сума"])
        ws3.cell(row=row, column=4, value=rr["Върнати бройки"])
        zebra_row(ws3, row, len(headers3))
        row += 1
    if row > 2:
        ws3.cell(row=row, column=1, value="ОБЩО").font = BOLD
        ws3.cell(row=row, column=3, value=f"=SUM(C2:C{row-1})").font = BOLD
        ws3.cell(row=row, column=4, value=f"=SUM(D2:D{row-1})").font = BOLD

    # ---------- ЛИСТ 4: КОНСИГНАЦИЯ ----------
    ws4 = wb.create_sheet("Консигнация")
    headers4 = ["Издателство", "Продадени бройки", "Сума за отчитане (без ДДС)",
                "Марж книжарница"]
    write_header(ws4, headers4)
    consign = get_consignment_report(date_from, date_to)
    row = 2
    for c in consign:
        ws4.cell(row=row, column=1, value=c["supplier_name"])
        ws4.cell(row=row, column=2, value=c["sold_qty"])
        ws4.cell(row=row, column=3, value=c["owed_to_publisher"])
        ws4.cell(row=row, column=4, value=c["bookstore_margin"])
        zebra_row(ws4, row, len(headers4))
        row += 1
    if row > 2:
        ws4.cell(row=row, column=1, value="ОБЩО").font = BOLD
        for col in (2, 3, 4):
            letter = chr(ord('A') + col - 1)
            ws4.cell(row=row, column=col,
                     value=f"=SUM({letter}2:{letter}{row-1})").font = BOLD

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def get_consignment_report(date_from, date_to):
    """
    Отчет за продадена КОНСИГНАЦИЯ, групиран по издателство.
    Правило: консигнацията се продава ПЪРВА. Затова за всяка книга
    продадените бройки се отчитат като консигнационни до размера на
    реално доставените консигнационни бройки (не повече).

    Работи на ниво "обща бройка за книга", което е достатъчно за отчета
    към издателството (колко им дължим за периода).
    """
    conn = get_connection()

    # За всяка книга вадим:
    #  - консигнационни доставени общо (за всички време),
    #  - продадени ПРЕДИ периода (за да знаем колко консигнация вече е изядена),
    #  - продадени В периода, с количествено-претеглени суми доставна/продажна.
    # Правилото "консигнация се продава първа" се прилага през ВРЕМЕТО, не само
    # вътре в периода — иначе по-стари продажби биха се отчели повторно.
    rows = conn.execute(
        """
        WITH
        consigned AS (
            SELECT product_id, SUM(quantity) AS consigned_qty
            FROM delivery_items
            WHERE settlement_type = 'Консигнация'
            GROUP BY product_id
        ),
        sold_before AS (
            SELECT si.product_id, SUM(si.quantity) AS qty
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            WHERE s.status != 'Отказана'
              AND si.product_id IS NOT NULL
              AND date(s.created_at) < ?
            GROUP BY si.product_id
        ),
        sold_period AS (
            SELECT si.product_id,
                   SUM(si.quantity) AS qty,
                   -- количествено претеглени суми (не AVG на единичните цени!)
                   SUM(si.quantity * si.cost_price) AS cost_val,
                   SUM(si.quantity * si.sale_price) AS sale_val
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            WHERE s.status != 'Отказана'
              AND si.product_id IS NOT NULL
              AND date(s.created_at) >= ? AND date(s.created_at) <= ?
            GROUP BY si.product_id
        )
        SELECT
            sup.name AS supplier_name,
            c.consigned_qty,
            COALESCE(sb.qty, 0) AS sold_before,
            sp.qty            AS sold_period_qty,
            sp.cost_val,
            sp.sale_val
        FROM sold_period sp
        JOIN consigned c   ON c.product_id = sp.product_id
        LEFT JOIN sold_before sb ON sb.product_id = sp.product_id
        JOIN products p    ON p.id = sp.product_id
        JOIN suppliers sup ON sup.id = p.supplier_id
        """,
        (date_from, date_from, date_to)
    ).fetchall()
    conn.close()

    # Прилагаме "консигнация първа" през времето и групираме по издателство.
    by_supplier = {}
    for r in rows:
        consigned = r["consigned_qty"]
        # Колко консигнация вече е изядена от продажбите ПРЕДИ периода.
        consumed_before = min(r["sold_before"], consigned)
        remaining_at_start = max(0, consigned - consumed_before)
        # Консигнационни бройки, продадени В периода (не повече от останалите).
        consign_sold = min(r["sold_period_qty"], remaining_at_start)
        if consign_sold <= 0:
            continue

        # Количествено претеглени единични цени за периода.
        period_qty = r["sold_period_qty"]
        unit_cost = r["cost_val"] / period_qty if period_qty else 0
        unit_sale = r["sale_val"] / period_qty if period_qty else 0

        agg = by_supplier.setdefault(
            r["supplier_name"],
            {"supplier_name": r["supplier_name"], "sold_qty": 0,
             "owed_to_publisher": 0.0, "bookstore_margin": 0.0})
        agg["sold_qty"] += consign_sold
        agg["owed_to_publisher"] += consign_sold * unit_cost
        agg["bookstore_margin"] += consign_sold * (unit_sale - unit_cost)

    result = sorted(by_supplier.values(), key=lambda a: a["supplier_name"])
    for a in result:
        a["owed_to_publisher"] = round(a["owed_to_publisher"], 2)
        a["bookstore_margin"] = round(a["bookstore_margin"], 2)
    return result

# ---------- ГЕНЕРИРАНЕ НА EXCEL ДНЕВНИК ----------

def build_monthly_payment_excel(year_month):
    """Excel с три листа: 'Неплатени', 'Платени', 'Оперативни Разходи' за месеца."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from io import BytesIO
    from db.expenses import get_expenses_by_period

    data = get_monthly_payment_report(year_month)
    wb = Workbook()

    # --- ЛИСТ 1: НЕПЛАТЕНИ ---
    ws = wb.active
    ws.title = "Неплатени"
    ws.append(["Дата/Час", "Поръчка №", "Товарителница",
               "Начин на плащане", "Сума"])
    for c in ws[1]:
        c.font = Font(bold=True)
    row = 2
    for r in data["unpaid"]:
        ws.append([
            r["created_at"],
            r["order_number"] or "-",
            r["waybill_number"] or "-",
            r["payment_method"],
            round(r["amount"], 2),
        ])
        row += 1
    if row > 2:
        ws.cell(row=row, column=1, value="ОБЩО ВИСЯЩИ").font = Font(bold=True)
        ws.cell(row=row, column=5, value=f"=SUM(E2:E{row-1})").font = Font(bold=True)

    # --- ЛИСТ 2: ПЛАТЕНИ ---
    ws2 = wb.create_sheet("Платени")
    ws2.append(["Дата/Час", "Поръчка №", "Товарителница",
                "Начин на плащане", "Сума", "Дата на плащане"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    row = 2
    for r in data["paid"]:
        ws2.append([
            r["created_at"],
            r["order_number"] or "-",
            r["waybill_number"] or "-",
            r["payment_method"],
            round(r["amount"], 2),
            r["payment_date"],
        ])
        row += 1
    if row > 2:
        ws2.cell(row=row, column=1, value="ОБЩО СЪБРАНИ").font = Font(bold=True)
        ws2.cell(row=row, column=5, value=f"=SUM(E2:E{row-1})").font = Font(bold=True)

    # --- ЛИСТ 3: ОПЕРАТИВНИ РАЗХОДИ ---
    # Групирани по категория, после по дата в рамките на категорията.
    # Между категориите вкарваме междинна сума, за да види счетоводителят
    # колко общо е отишло за наем, колко за заплати и т.н.
    ws3 = wb.create_sheet("Оперативни Разходи")
    ws3.append(["Дата", "Категория", "Описание", "Сума", "Документ №"])
    for c in ws3[1]:
        c.font = Font(bold=True)

    # Изчисляваме границите на месеца от 'year_month' формат 'YYYY-MM'.
    year, month = year_month.split("-")
    date_from = f"{year}-{month}-01"
    # Последен ден на месеца — взимаме първия на следващия и махаме един ден.
    from datetime import date as _date, timedelta as _timedelta
    next_month = int(month) + 1
    next_year = int(year)
    if next_month == 13:
        next_month = 1
        next_year += 1
    date_to = (_date(next_year, next_month, 1) - _timedelta(days=1)).isoformat()

    expenses = get_expenses_by_period(date_from, date_to)

    # Групираме по категория. expenses вече идва подреден по date DESC,
    # но за листа предпочитаме категория-после-дата, затова сортираме наново.
    expenses_sorted = sorted(expenses, key=lambda e: (e["category"], e["date"]))

    row = 2
    first_row_of_category = row
    current_category = None
    grand_total_start = row

    for e in expenses_sorted:
        # Когато сменим категория, добавяме междинна сума за предишната.
        if current_category is not None and e["category"] != current_category:
            ws3.cell(row=row, column=2,
                     value=f"Общо: {current_category}").font = Font(bold=True, italic=True)
            ws3.cell(row=row, column=4,
                     value=f"=SUM(D{first_row_of_category}:D{row-1})"
                     ).font = Font(bold=True, italic=True)
            row += 1
            first_row_of_category = row

        ws3.append([
            e["date"],
            e["category"],
            e["description"] or "-",
            round(e["amount"], 2),
            e["document_number"] or "-",
        ])
        current_category = e["category"]
        row += 1

    # След последния запис, междинна сума за последната категория.
    if current_category is not None:
        ws3.cell(row=row, column=2,
                 value=f"Общо: {current_category}").font = Font(bold=True, italic=True)
        ws3.cell(row=row, column=4,
                 value=f"=SUM(D{first_row_of_category}:D{row-1})"
                 ).font = Font(bold=True, italic=True)
        row += 1

    # Финален общ сбор — само за реалните данни, не за междинните суми.
    # Затова сумираме поредицата от D2:D{последен_ред}, но с филтър.
    # По-просто решение: записваме директна формула SUMIF за категория != "Общо:..."
    # Или: пазим списък от редове, които са истински данни.
    # За простота — сумираме директните стойности от Python (не формула).
    if expenses_sorted:
        total = sum(e["amount"] for e in expenses_sorted)
        ws3.cell(row=row+1, column=2, value="ОБЩО ЗА МЕСЕЦА").font = Font(bold=True, size=12)
        ws3.cell(row=row+1, column=4, value=round(total, 2)).font = Font(bold=True, size=12)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()  


def build_accounting_excel(date_from, date_to):
    """
    Връща Excel файл (в паметта, като bytes) с два листа:
    'Продажби' и 'Покупки'. Сумите с ДДС са стойности; основата и ДДС
    са Excel ФОРМУЛИ, за да е документът проверим и преизчислим.
    Връща bytes, готови за st.download_button.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from io import BytesIO

    wb = Workbook()

    # --- ЛИСТ 1: ПРОДАЖБИ ---
    ws = wb.active
    ws.title = "Продажби"
    headers = ["Документ", "Данъчна основа", "ДДС ставка", "Начислено ДДС",
               "Обща стойност с ДДС", "Фискална група", "Тип плащане", "Статус"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    sales = get_sales_journal(date_from, date_to)
    row = 2
    for s in sales:
        total = s["Обща стойност с ДДС"]
        rate = s["ДДС ставка"]
        ws.cell(row=row, column=1, value=s["Документ"])
        if rate == "0%":
            ws.cell(row=row, column=2, value=total)
            ws.cell(row=row, column=3, value="0%")
            ws.cell(row=row, column=4, value=0)
        else:
            # Правилният делител според ставката (1.09 за 9%, 1.20 за 20%).
            divisor = "1.20" if rate == "20%" else "1.09"
            ws.cell(row=row, column=2, value=f"=E{row}/{divisor}")
            ws.cell(row=row, column=3, value=rate)
            ws.cell(row=row, column=4, value=f"=E{row}-B{row}")
        ws.cell(row=row, column=5, value=total)
        ws.cell(row=row, column=6, value=s["Фискална група"])
        ws.cell(row=row, column=7, value=s["Тип плащане"])
        ws.cell(row=row, column=8, value=s["Статус"])
        row += 1

    if row > 2:
        ws.cell(row=row, column=1, value="ОБЩО").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"=SUM(B2:B{row-1})").font = Font(bold=True)
        ws.cell(row=row, column=4, value=f"=SUM(D2:D{row-1})").font = Font(bold=True)
        ws.cell(row=row, column=5, value=f"=SUM(E2:E{row-1})").font = Font(bold=True)

    # --- ЛИСТ 2: ПОКУПКИ ---
    ws2 = wb.create_sheet("Покупки")
    headers2 = ["Фактура доставчик", "Доставчик", "Данъчна основа",
                "ДДС 9%", "Общо платена сума"]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    purchases = get_purchases_journal(date_from, date_to)
    row = 2
    for p in purchases:
        total = p["Общо платена сума"]
        ws2.cell(row=row, column=1, value=p["Фактура доставчик"])
        ws2.cell(row=row, column=2, value=p["Доставчик"])
        ws2.cell(row=row, column=3, value=f"=E{row}/1.09")
        ws2.cell(row=row, column=4, value=f"=E{row}-C{row}")
        ws2.cell(row=row, column=5, value=total)
        row += 1

    if row > 2:
        ws2.cell(row=row, column=1, value="ОБЩО").font = Font(bold=True)
        ws2.cell(row=row, column=3, value=f"=SUM(C2:C{row-1})").font = Font(bold=True)
        ws2.cell(row=row, column=4, value=f"=SUM(D2:D{row-1})").font = Font(bold=True)
        ws2.cell(row=row, column=5, value=f"=SUM(E2:E{row-1})").font = Font(bold=True)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()    



# ---------- МЕСЕЧЕН ОТЧЕТ ЗА СТАТУС НА ПЛАЩАНИЯТА ----------

def get_monthly_payment_report(year_month):
    """
    Връща месечен отчет за плащанията. year_month е низ 'YYYY-MM'.
    Дели поръчките на неплатени (чакащи) и платени, за месеца по дата на създаване.
    Връща речник с трите суми и двата списъка.
    """
    conn = get_connection()

    # Неплатени (чакащи плащане) за месеца
    unpaid = conn.execute(
        """SELECT s.created_at, s.order_number, s.waybill_number,
                  s.payment_method,
                  s.supplementary_payment_method,
                  s.supplementary_amount,
                  COALESCE(SUM(si.quantity * si.sale_price), 0) AS amount
           FROM sales s
           LEFT JOIN sale_items si ON si.sale_id = s.id
           WHERE s.status = 'Чака плащане'
             AND strftime('%Y-%m', s.created_at) = ?
           GROUP BY s.id
           ORDER BY s.created_at""",
        (year_month,)
    ).fetchall()

    # Платени за месеца (с дата на плащане)
    paid = conn.execute(
        """SELECT s.created_at, s.order_number, s.waybill_number,
                  s.payment_method, 
                  s.supplementary_payment_method,
                  s.supplementary_amount,
                  s.payment_date,
                  COALESCE(SUM(si.quantity * si.sale_price), 0) AS amount
           FROM sales s
           LEFT JOIN sale_items si ON si.sale_id = s.id
           WHERE s.status = 'Платена'
             AND strftime('%Y-%m', s.created_at) = ?
           GROUP BY s.id
           ORDER BY s.created_at""",
        (year_month,)
    ).fetchall()

    unpaid_total = sum(r["amount"] for r in unpaid)
    paid_total = sum(r["amount"] for r in paid)

    def describe(r):
        d = dict(r)
        if d["payment_method"] == "Ваучер" and d["supplementary_amount"] > 0:
            d["payment_method"] = (f"Ваучер + {d['supplementary_payment_method']} "
                                   f"({d['supplementary_amount']:.2f} лв.)")
        return d

    return {
        "unpaid": [describe(r) for r in unpaid],
        "paid": [describe(r) for r in paid],
        "unpaid_total": unpaid_total,
        "paid_total": paid_total,
        "turnover": unpaid_total + paid_total,
    }

