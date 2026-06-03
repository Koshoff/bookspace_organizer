from db.connection import get_connection
# само АКО функцията хваща sqlite3.IntegrityError:
import sqlite3


# ---------- ГЛАВНО КОМАНДНО ТАБЛО (Модул 0) ----------

def get_dashboard_data(date_from, date_to, payment_method=None):
    """
    Връща всички данни за командното табло за зададения период.
    date_from / date_to са низове 'YYYY-MM-DD'.

    ВАЖНО за филтрирането по дата:
    - Приходите (платени продажби) се филтрират по payment_date (кога е платено).
    - Разходите (платени доставки) се филтрират по delivery_paid_date (кога сме платили).
    Така и двете страни гледат КОГА реално са минали парите.
    """
    conn = get_connection()

    # --- КАРТА 1: Приходи = платени продажби, по дата на плащане ---
    rev_query = """SELECT COALESCE(SUM(si.quantity * si.sale_price), 0) AS total
                   FROM sales s
                   JOIN sale_items si ON si.sale_id = s.id
                   WHERE s.status = 'Платена'
                     AND date(s.payment_date) >= ? AND date(s.payment_date) <= ?"""
    rev_params = [date_from, date_to]
    if payment_method:
        rev_query += " AND s.payment_method = ?"
        rev_params.append(payment_method)
    revenue = conn.execute(rev_query, rev_params).fetchone()["total"]

    # --- КАРТА 2: Разходи = платени доставки, по дата на плащане ---
    expenses = conn.execute(
        """SELECT COALESCE(SUM(di.quantity * di.delivery_price), 0) AS total
           FROM deliveries d
           JOIN delivery_items di ON di.delivery_id = d.id
           WHERE d.payment_status = 'Платена'
             AND date(d.delivery_paid_date) >= ? AND date(d.delivery_paid_date) <= ?""",
        (date_from, date_to)
    ).fetchone()["total"]

    # --- КАРТА 4: Брой продажби за периода (по дата на създаване, не отказани) ---
    sales_count = conn.execute(
        """SELECT COUNT(*) AS c
           FROM sales
           WHERE date(created_at) >= ? AND date(created_at) <= ?
             AND status != 'Отказана'""",
        (date_from, date_to)
    ).fetchone()["c"]

    # --- ЛЯВА КОЛОНА: Задължения — неплатени доставки (най-новите отгоре) ---
    # Тези НЕ се филтрират по период — дължиш ги СЕГА, независимо кога са вкарани.
    liabilities = conn.execute(
        """SELECT d.id, d.doc_type, d.doc_number, d.doc_date, s.name AS supplier_name,
                  COALESCE(SUM(di.quantity * di.delivery_price), 0) AS amount
           FROM deliveries d
           JOIN suppliers s ON s.id = d.supplier_id
           LEFT JOIN delivery_items di ON di.delivery_id = d.id
           WHERE d.payment_status = 'Неплатена'
           GROUP BY d.id
           ORDER BY d.doc_date DESC"""
    ).fetchall()

    # --- ДЯСНА КОЛОНА: Вземания — продажби, чакащи плащане (наложен платеж) ---
    receivables = conn.execute(
        """SELECT s.id, s.order_number, s.waybill_number, s.created_at,
                  COALESCE(SUM(si.quantity * si.sale_price), 0) AS amount
           FROM sales s
           LEFT JOIN sale_items si ON si.sale_id = s.id
           WHERE s.status = 'Чака плащане'
           GROUP BY s.id
           ORDER BY s.created_at DESC"""
    ).fetchall()

    # --- ДОЛЕН ПАНЕЛ: последни 10 активности (смесено от трите вида) ---
    # UNION ALL слепва три заявки в общ списък. Всяка дава еднакви колони:
    # дата, тип, документ, стойност — за да се подредят заедно по време.
    activities = conn.execute(
        """SELECT * FROM (
               -- Продажби
               SELECT s.created_at AS ts,
                      CASE
                          WHEN s.order_number LIKE 'VOUCHER-%' THEN 'Издаване ваучер'
                          WHEN s.payment_method = 'Ваучер' THEN 'Продажба (с ваучер)'
                          ELSE 'Продажба'
                      END AS type,
                      'Поръчка №' || COALESCE(s.order_number, '-') AS doc,
                      COALESCE((SELECT SUM(si.quantity * si.sale_price)
                                FROM sale_items si WHERE si.sale_id = s.id), 0) AS value
               FROM sales s

               UNION ALL

               -- Доставки
               SELECT d.created_at AS ts, 'Доставка' AS type,
                      d.doc_type || ' №' || d.doc_number AS doc,
                      COALESCE((SELECT SUM(di.quantity * di.delivery_price)
                                FROM delivery_items di WHERE di.delivery_id = d.id), 0) AS value
               FROM deliveries d

               UNION ALL

               -- Кредитни известия
               SELECT cn.created_at AS ts, 'Кредитно известие' AS type,
                      'Бележка №' || cn.original_receipt AS doc,
                      cn.returned_amount AS value
               FROM credit_notes cn
           )
           ORDER BY ts DESC
           LIMIT 10"""
    ).fetchall()

    conn.close()

    return {
        "revenue": revenue,
        "expenses": expenses,
        "profit": revenue - expenses,        # Карта 3: чиста печалба
        "sales_count": sales_count,
        "liabilities": liabilities,
        "receivables": receivables,
        "activities": activities,
    }



# ---------- МЕСЕЧЕН ОТЧЕТ ЗА СТАТУС НА ПЛАЩАНИЯТА ----------

def get_monthly_payment_report(year_month):
    """
    Връща месечен отчет за плащанията. year_month е низ 'YYYY-MM'.
    Дели поръчките на неплатени (чакащи) и платени, за месеца по дата на създаване.
    Връща речник с трите суми и двата списъка.
    """
    conn = get_connection()

    # Неплатени (чакащи плащане) за месеца
    unpaid = conn.execute(
        """SELECT s.created_at, s.order_number, s.waybill_number,
                  s.payment_method,
                  s.supplementary_payment_method,
                  s.supplementary_amount,
                  COALESCE(SUM(si.quantity * si.sale_price), 0) AS amount
           FROM sales s
           LEFT JOIN sale_items si ON si.sale_id = s.id
           WHERE s.status = 'Чака плащане'
             AND strftime('%Y-%m', s.created_at) = ?
           GROUP BY s.id
           ORDER BY s.created_at""",
        (year_month,)
    ).fetchall()

    # Платени за месеца (с дата на плащане)
    paid = conn.execute(
        """SELECT s.created_at, s.order_number, s.waybill_number,
                  s.payment_method, 
                  s.supplementary_payment_method,
                  s.supplementary_amount,
                  s.payment_date,
                  COALESCE(SUM(si.quantity * si.sale_price), 0) AS amount
           FROM sales s
           LEFT JOIN sale_items si ON si.sale_id = s.id
           WHERE s.status = 'Платена'
             AND strftime('%Y-%m', s.created_at) = ?
           GROUP BY s.id
           ORDER BY s.created_at""",
        (year_month,)
    ).fetchall()

    unpaid_total = sum(r["amount"] for r in unpaid)
    paid_total = sum(r["amount"] for r in paid)

    def describe(r):
        d = dict(r)
        if d["payment_method"] == "Ваучер" and d["supplementary_amount"] > 0:
            d["payment_method"] = (f"Ваучер + {d['supplementary_payment_method']} "
                                   f"({d['supplementary_amount']:.2f} лв.)")
        return d

    return {
        "unpaid": [describe(r) for r in unpaid],
        "paid": [describe(r) for r in paid],
        "unpaid_total": unpaid_total,
        "paid_total": paid_total,
        "turnover": unpaid_total + paid_total,
    }

def build_monthly_payment_excel(year_month):
    """Excel с два листа: 'Неплатени' и 'Платени' за месеца."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from io import BytesIO

    data = get_monthly_payment_report(year_month)
    wb = Workbook()

    # Лист 1: Неплатени
    ws = wb.active
    ws.title = "Неплатени"
    ws.append(["Дата/Час", "Поръчка №", "Товарителница",
               "Начин на плащане", "Сума"])
    for c in ws[1]:
        c.font = Font(bold=True)
    row = 2
    for r in data["unpaid"]:
        ws.cell(row=row, column=1, value=r["created_at"])
        ws.cell(row=row, column=2, value=r["order_number"] or "-")
        ws.cell(row=row, column=3, value=r["waybill_number"] or "-")
        ws.cell(row=row, column=4, value=r["payment_method"])
        ws.cell(row=row, column=5, value=round(r["amount"], 2))
        row += 1
    if row > 2:
        ws.cell(row=row, column=1, value="ОБЩО ВИСЯЩИ").font = Font(bold=True)
        ws.cell(row=row, column=5, value=f"=SUM(E2:E{row-1})").font = Font(bold=True)

    # Лист 2: Платени
    ws2 = wb.create_sheet("Платени")
    ws2.append(["Дата/Час", "Поръчка №", "Товарителница",
                "Начин на плащане", "Сума", "Дата на плащане"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    row = 2
    for r in data["paid"]:
        ws2.cell(row=row, column=1, value=r["created_at"])
        ws2.cell(row=row, column=2, value=r["order_number"] or "-")
        ws2.cell(row=row, column=3, value=r["waybill_number"] or "-")
        ws2.cell(row=row, column=4, value=r["payment_method"])
        ws2.cell(row=row, column=5, value=round(r["amount"], 2))
        ws2.cell(row=row, column=6, value=r["payment_date"])
        row += 1
    if row > 2:
        ws2.cell(row=row, column=1, value="ОБЩО СЪБРАНИ").font = Font(bold=True)
        ws2.cell(row=row, column=5, value=f"=SUM(E2:E{row-1})").font = Font(bold=True)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()




